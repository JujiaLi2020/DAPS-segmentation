r"""
DAPS Calibration Lab.

Launch from the project root:
    .\.venv\Scripts\python.exe code\daps_calibration_lab.py

This tool is intentionally separate from the segmentation UI. It creates
multi-annotator boundary-labeling templates and evaluates completed labels.
"""

from __future__ import annotations

from collections import Counter
from datetime import datetime
import html
import json
import os
from pathlib import Path
import re
import sqlite3
from typing import Iterable

import gradio as gr
import numpy as np
import pandas as pd

from daps_excel_segmenter import (
    DAPSConfig,
    DAPSSegmenter,
    SPEAKER_LABEL_RE,
    SimilarityModel,
    SignalVocabulary,
    clean_dataframe_text,
    infer_id_column,
    infer_text_columns,
    normalize_transcript,
    read_table,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]


def resolve_output_dir() -> Path:
    configured_dir = os.environ.get("DAPS_OUTPUT_DIR") or os.environ.get("DAPS_STORAGE_DIR")
    if configured_dir:
        return Path(configured_dir)

    railway_volume = Path("/data")
    is_railway = any(
        os.environ.get(name)
        for name in (
            "RAILWAY_ENVIRONMENT",
            "RAILWAY_ENVIRONMENT_NAME",
            "RAILWAY_PROJECT_ID",
            "RAILWAY_SERVICE_ID",
        )
    )
    if is_railway and railway_volume.exists():
        return railway_volume

    return PROJECT_ROOT / "data" / "calibration_outputs"


OUTPUT_DIR = resolve_output_dir()
DB_PATH = OUTPUT_DIR / "daps_calibration.sqlite3"
DEFAULT_VOCABULARY = SignalVocabulary()
SIGNALS = ["C_t", "M_t", "A_t", "R_t"]
SIGNAL_COLUMNS = {
    "C_t": "cognitive_transition",
    "M_t": "metacognitive_reset",
    "A_t": "affective_friction",
    "R_t": "structural_break",
}
TOKEN_RE = re.compile(r"[A-Za-z]+(?:['’][A-Za-z]+)?")

APP_CSS = """
.step-title {
    display: flex;
    align-items: center;
    gap: 8px;
    font-size: 18px;
    font-weight: 700;
    margin: 8px 0 2px;
}
.help-dot {
    position: relative;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 17px;
    height: 17px;
    border: 1px solid var(--border-color-primary);
    border-radius: 50%;
    color: var(--body-text-color-subdued);
    font-size: 11px;
    cursor: help;
}
.help-dot:hover::after {
    content: attr(data-tooltip);
    position: absolute;
    z-index: 100;
    left: 50%;
    bottom: 130%;
    transform: translateX(-50%);
    width: 380px;
    padding: 10px 12px;
    border: 1px solid var(--border-color-primary);
    border-radius: 6px;
    background: var(--background-fill-primary);
    color: var(--body-text-color);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.22);
    font-size: 12px;
    font-weight: 400;
    line-height: 1.35;
    white-space: normal;
}
.step-note {
    color: var(--body-text-color-subdued);
    font-size: 13px;
    margin: 0 0 8px;
}
.status-table {
    width: 100%;
    border-collapse: collapse;
    margin: 8px 0 12px;
    font-size: 13px;
}
.status-table th,
.status-table td {
    border: 1px solid var(--border-color-primary);
    padding: 7px 9px;
    text-align: left;
    vertical-align: top;
}
.status-table th {
    background: var(--background-fill-secondary);
    font-weight: 700;
}
.status-pass td {
    background: rgba(35, 134, 54, 0.15);
}
.status-warning td {
    background: rgba(187, 128, 9, 0.18);
}
.status-fail td {
    background: rgba(218, 54, 51, 0.18);
}
.status-info td {
    background: rgba(56, 139, 253, 0.12);
}
.status-badge {
    display: inline-block;
    min-width: 72px;
    padding: 2px 7px;
    border-radius: 999px;
    font-size: 12px;
    font-weight: 700;
    text-align: center;
}
.status-pass .status-badge {
    color: #2ea043;
    border: 1px solid rgba(46, 160, 67, 0.6);
}
.status-warning .status-badge {
    color: #d29922;
    border: 1px solid rgba(210, 153, 34, 0.7);
}
.status-fail .status-badge {
    color: #f85149;
    border: 1px solid rgba(248, 81, 73, 0.7);
}
.status-info .status-badge {
    color: #58a6ff;
    border: 1px solid rgba(88, 166, 255, 0.6);
}
.metric-help-dot {
    position: relative;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 15px;
    height: 15px;
    margin-left: 5px;
    border: 1px solid currentColor;
    border-radius: 50%;
    color: var(--body-text-color-subdued);
    font-size: 10px;
    cursor: help;
}
.metric-help-dot:hover::after {
    content: attr(data-tooltip);
    position: absolute;
    z-index: 200;
    left: 0;
    bottom: 135%;
    width: 360px;
    padding: 9px 11px;
    border: 1px solid var(--border-color-primary);
    border-radius: 6px;
    background: var(--background-fill-primary);
    color: var(--body-text-color);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.22);
    font-size: 12px;
    font-weight: 400;
    line-height: 1.35;
    white-space: normal;
}
.field-label {
    display: flex;
    align-items: center;
    gap: 6px;
    font-weight: 600;
    font-size: 13px;
    margin: 0 0 4px;
}
.parameter-row {
    align-items: end;
    gap: 14px;
}
.parameter-field {
    min-width: 0;
}
.parameter-field .wrap {
    gap: 4px;
}
.parameter-symbol {
    color: var(--body-text-color-subdued);
    font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    font-size: 12px;
}
.review-layout-note {
    color: var(--body-text-color-subdued);
    font-size: 13px;
}
.review-context-box {
    border: 1px solid var(--border-color-primary);
    border-radius: 8px;
    padding: 14px;
    background: var(--background-fill-secondary);
    line-height: 1.55;
    font-size: 15px;
}
.review-boundary-marker {
    display: inline-block;
    padding: 2px 8px;
    margin: 0 3px;
    border-radius: 999px;
    background: rgba(255, 128, 0, 0.22);
    border: 1px solid rgba(255, 128, 0, 0.75);
    color: #ffb86b;
    font-weight: 700;
}
.review-segment-grid {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 10px;
    margin-top: 10px;
}
.review-segment-card {
    border: 1px solid var(--border-color-primary);
    border-radius: 8px;
    padding: 10px;
    background: var(--background-fill-primary);
    min-height: 92px;
}
.review-segment-card h4 {
    margin: 0 0 6px;
    font-size: 12px;
    color: var(--body-text-color-subdued);
}
.review-progress {
    margin: 8px 0;
    padding: 10px 12px;
    border: 1px solid var(--border-color-primary);
    border-radius: 8px;
    background: var(--background-fill-secondary);
}
.review-help {
    font-size: 12px;
    color: var(--body-text-color-subdued);
    margin-top: -6px;
    margin-bottom: 6px;
}
"""

STEP_HELP = {
    "Step 0. Segment Raw Transcript Data": (
        "Start here when you have raw CSV/Excel transcript data rather than a DAPS workbook. "
        "For the current input.csv structure, use ID as the participant column, Item as the item/task column, "
        "and Transcript as the text column. This step reads Windows-1252 CSV files when UTF-8 fails, cleans mojibake artifacts, "
        "runs DAPS segmentation, and writes a workbook with segments and boundaries sheets."
    ),
    "Step 1. Inspect Segmentation Workbook": (
        "Load a DAPS output workbook and run pre-validation checks before annotation. "
        "This step verifies that the workbook has segments and boundaries sheets, summarizes segment length, "
        "checks event/speaker leakage, and records whether the output is suitable for calibration sampling."
    ),
    "Step 2. Generate Multi-Annotator Template": (
        "Create a boundary-context annotation template. Pilot mode is for approximately 50 items to test the codebook; "
        "formal mode is for 400-500 items. The same sampled boundary items are duplicated for each annotator so agreement can be calculated."
    ),
    "Step 2b. In-App Human Review": (
        "Review annotation items directly in the app instead of editing a wide spreadsheet. "
        "The app shows context around the candidate boundary, provides guided controls for each human label, "
        "saves every decision to SQLite, and can export a reviewed workbook for agreement analysis."
    ),
    "Step 3. Evaluate Multi-Annotator Labels": (
        "Upload completed annotation files and compute label completion, positive rates, exact agreement, and two-annotator Cohen's kappa. "
        "Use this step to identify unclear signal definitions before training vocabulary lists."
    ),
    "Step 4. Analyze Cue Lexicon": (
        "Estimate cue precision, recall, support, and lift from completed human labels. This is the first data-driven step toward vocabulary training. "
        "Candidate words and phrases are recommended only when they have sufficient support and precision."
    ),
    "Step 5. SQLite History": (
        "Every major operation is recorded in a local SQLite database with timestamp, operation type, input file, parameters, status, and output path. "
        "Use this tab to audit calibration decisions and reproduce a run."
    ),
}

METRIC_HELP = {
    "Segments": "Total number of output DAPS segments across all source records. This is descriptive and should be reported, not passed or failed by itself.",
    "Boundaries": "Total number of token gaps evaluated as possible boundaries in the boundaries sheet.",
    "Records": "Number of unique Record_ID values represented in the segments sheet.",
    "Median segment tokens": "Median segment length in tokens. Suggested range: 12-25 tokens; much lower suggests oversegmentation.",
    "Mean segment tokens": "Average segment length in tokens. Suggested range: 15-30 tokens; this complements the median.",
    "Short segments < 8 tokens": "Raw count of segments shorter than 8 tokens. Use the short segment rate for the main quality judgment.",
    "Short segment rate": "Proportion of segments shorter than 8 tokens. Values above .25 suggest too many tiny units.",
    "Long segments > 60 tokens": "Count of segments longer than the default L_max=60 target. Nonzero values suggest missed process shifts or forced-split issues.",
    "Max tokens": "Length of the longest segment. It should usually stay within the configured L_max.",
    "Max segments per record": "Largest number of segments produced from a single Record_ID. Very high values suggest record-level oversegmentation.",
    "Event token leakage": "Counts internal event placeholders such as <EVENT> that accidentally remained in Segment_Text. These should be removed or separated.",
    "Bracket event leakage": "Counts bracketed transcript events such as [Pause], [Drawing], [End of Audio], or [Unintelligible] that remain inside ordinary segment text.",
    "Speaker label leakage": "Counts speaker labels such as Interviewer:, Interviewee:, Participant:, Child:, or Speaker A: that remain inside segment text. These labels should usually be metadata or separate events, not part of the reasoning segment.",
    "Punctuation-leading segments": "Counts segments that begin with punctuation such as comma or period. Punctuation should usually attach to the previous segment.",
    "Selected boundaries": "Number of boundaries selected by the DAPS decoder as final cuts.",
    "Mean semantic gravity: selected": "Average G_t for selected boundaries. In this implementation, selected boundaries should usually have lower semantic gravity than non-selected gaps.",
    "Mean semantic gravity: non-selected": "Average G_t for gaps not selected as final boundaries. This should usually be higher than selected-boundary G_t.",
    "Mean transition pressure: selected": "Average transition pressure for selected boundaries. Selected cuts should usually have higher pressure than non-selected gaps.",
    "Mean transition pressure: non-selected": "Average transition pressure for non-selected gaps. This should usually be lower than selected-boundary pressure.",
}

SIGNAL_HELP = {
    "C_t": "Cognitive transition: evidence of a shift in object, strategy, operation, or reasoning state.",
    "M_t": "Metacognitive reset: evidence of monitoring, uncertainty, correction, re-planning, or realization.",
    "A_t": "Affective friction: evidence of frustration, difficulty, confidence change, or emotional/effortful resistance.",
    "R_t": "Rhetorical/structural break: evidence of discourse structure, speaker turn, event marker, pause, or utterance organization.",
}

ANNOTATION_INSTRUCTIONS = [
    {
        "Section": "What you are annotating",
        "Instruction": "Each row is one candidate boundary. In Boundary_Context, [[CANDIDATE_BOUNDARY]] marks the candidate break being judged. It is not transcript text.",
        "Example": "I tried rotating it. [[CANDIDATE_BOUNDARY]] Now I flip it over.",
    },
    {
        "Section": "Read these columns first",
        "Instruction": "Read Previous_Segment, Left_Segment, Right_Segment, and Boundary_Context before filling labels. Treat the target boundary as the break between Left_Segment and Right_Segment.",
        "Example": "Use Boundary_Context for local wording; use Left_Segment and Right_Segment to judge whether the split is meaningful.",
    },
    {
        "Section": "Do not edit metadata",
        "Instruction": "Only fill human annotation columns. Do not edit Annotation_ID, Annotation_Item_ID, Record_ID, Boundary_Gap, context columns, algorithm score columns, or sampling columns.",
        "Example": "Fill Boundary_Strength_0_3, Human_* signal fields, Cue_Span, Codebook_Issue, and Notes only.",
    },
    {
        "Section": "Boundary_Strength_0_3",
        "Instruction": "Enter one 0-3 judgment for the candidate boundary: 0 = merge/not a boundary; 1 = weak or uncertain boundary; 2 = moderate/acceptable boundary; 3 = strong/clear boundary.",
        "Example": "3 for a clear shift from counting to explaining; 0 for this [[CANDIDATE_BOUNDARY]] one or a [[CANDIDATE_BOUNDARY]] right triangle.",
    },
    {
        "Section": "Human_C_t",
        "Instruction": "Enter 1 when there is a cognitive transition: object, shape, operation, strategy, comparison, exclusion, or reasoning state changes. Otherwise enter 0.",
        "Example": "I look at this one -> Then I compare it with that one; rotate -> flip; possible answer -> X out.",
    },
    {
        "Section": "Human_M_t",
        "Instruction": "Enter 1 when there is metacognitive monitoring/reset: uncertainty, self-correction, checking, realization, re-planning, or rejection of a prior thought. Otherwise enter 0.",
        "Example": "wait; actually; I don't know; maybe; now I know; that can't work. Single 'no' counts only when it corrects or rejects a prior thought.",
    },
    {
        "Section": "Human_A_t",
        "Instruction": "Enter 1 when there is affective friction: frustration, difficulty, confidence change, effort, or emotional resistance. Otherwise enter 0.",
        "Example": "hard; confusing; tricky; ugh; I can't; I'm stuck. Do not mark 'looks like' as affective when it only means visual comparison.",
    },
    {
        "Section": "Human_R_t",
        "Instruction": "Enter 1 when the boundary is rhetorical/structural: speaker turn, pause/event, discourse connector, new utterance organization, or task-transition marker. Otherwise enter 0.",
        "Example": "then; okay; next; so; but; [Pause]; [Drawing]; interviewer question -> participant answer.",
    },
    {
        "Section": "Multi-label rule",
        "Instruction": "Human_C_t, Human_M_t, Human_A_t, and Human_R_t are independent 0/1 labels. More than one can be 1 for the same boundary.",
        "Example": "I think it fits. [[CANDIDATE_BOUNDARY]] Wait, no, I need to flip it. This can be C_t=1 and M_t=1.",
    },
    {
        "Section": "Human_Primary_Type",
        "Instruction": "Choose the strongest single label: cognitive, metacognitive, affective, structural, mixed, low_signal, or unclear. Use mixed when two or more signals are equally central.",
        "Example": "Use low_signal when the boundary is valid but has weak signal evidence; use unclear when you cannot decide.",
    },
    {
        "Section": "Evidence log: one idea, three boxes",
        "Instruction": "After choosing labels, use the three evidence boxes as a simple audit trail. Box 1 records what supports your label, Box 2 records cue-like words you rejected, and Box 3 records rules that need team discussion.",
        "Example": "Most rows only need Box 1. Use Box 2 only when a misleading cue appears. Use Box 3 only when the codebook needs clarification.",
    },
    {
        "Section": "Evidence box 1: supporting cue",
        "Instruction": "Cue_Span is the supporting evidence box. Copy the exact word or phrase that made you mark one of Human_C_t, Human_M_t, Human_A_t, or Human_R_t as 1. If all four signal labels are 0, usually leave it blank.",
        "Example": "M_t=1 because of 'wait, no' -> Cue_Span: wait, no. C_t=1 because of 'flip it over' -> Cue_Span: flip it over. R_t=1 because of '[Pause]' -> Cue_Span: [Pause].",
    },
    {
        "Section": "Evidence box 2: rejected cue",
        "Instruction": "Counterexample_or_Exclusion is the rejected-cue box. Use it when a word/phrase may look like a cue, but you decided it should NOT count for a signal in this context. Leave blank if there is no misleading cue.",
        "Example": "'looks like' is visual comparison, not A_t. 'no' is an answer choice label, not M_t. 'hard' describes the object, not the child's difficulty.",
    },
    {
        "Section": "Evidence box 3: rule question",
        "Instruction": "Codebook_Issue is the rule-question box. Use it only when the annotation rule or vocabulary definition needs discussion. This is not for ordinary evidence; it flags cases the research team should review after pilot annotation.",
        "Example": "Unsure whether 'wait' should count as M_t or R_t here. Boundary seems valid but none of the four signal definitions fits. Need rule for interviewer prompt transitions.",
    },
    {
        "Section": "If boundary is invalid",
        "Instruction": "If Boundary_Strength_0_3 = 0 because the split is clearly wrong, usually set Human_C_t=0, Human_M_t=0, Human_A_t=0, Human_R_t=0, and Human_Primary_Type=low_signal or unclear.",
        "Example": "For this [[CANDIDATE_BOUNDARY]] one, mark Boundary_Strength_0_3 = 0.",
    },
]

ANNOTATION_CODEBOOK_ROWS = [
    {
        "Field": "Boundary_Strength_0_3",
        "Short label": "Boundary",
        "Required": "Yes",
        "Allowed values": "0/1/2/3",
        "How to fill": "0 = merge/not a boundary; 1 = weak boundary; 2 = moderate boundary; 3 = strong boundary.",
        "Example": "3 for a clear process shift; 0 for a phrase-internal split.",
    },
    {
        "Field": "Human_C_t",
        "Short label": "C_t",
        "Required": "Yes",
        "Allowed values": "0/1",
        "How to fill": "1 when there is a cognitive transition: object, operation, strategy, comparison, or reasoning state changes.",
        "Example": "rotate -> flip; try this -> compare that; same -> different.",
    },
    {
        "Field": "Human_M_t",
        "Short label": "M_t",
        "Required": "Yes",
        "Allowed values": "0/1",
        "How to fill": "1 when there is monitoring, correction, uncertainty, realization, checking, or re-planning.",
        "Example": "wait; actually; I don't know; now I know; maybe; that can't work.",
    },
    {
        "Field": "Human_A_t",
        "Short label": "A_t",
        "Required": "Yes",
        "Allowed values": "0/1",
        "How to fill": "1 when affective friction is present: frustration, difficulty, confidence/effort expression, or emotional resistance.",
        "Example": "this is hard; confusing; ugh; tricky. Do not mark 'looks like' as affective by itself.",
    },
    {
        "Field": "Human_R_t",
        "Short label": "R_t",
        "Required": "Yes",
        "Allowed values": "0/1",
        "How to fill": "1 when the boundary is mainly rhetorical or structural: speaker turn, event marker, pause, discourse connector, or utterance organization.",
        "Example": "then; okay; interviewer question -> participant answer; [Pause]; [Drawing].",
    },
    {
        "Field": "Human_Primary_Type",
        "Short label": "Primary",
        "Required": "Recommended",
        "Allowed values": "cognitive/metacognitive/affective/structural/mixed/low_signal/unclear",
        "How to fill": "Choose the strongest reason for the boundary. Use mixed if two or more signals are equally central.",
        "Example": "mixed when a boundary is both a strategy shift and a self-correction.",
    },
    {
        "Field": "Cue_Span",
        "Short label": "Cue",
        "Required": "Recommended if any signal = 1",
        "Allowed values": "short text",
        "How to fill": "Evidence box 1: supporting cue. Copy exact positive evidence for any signal label you marked as 1. Leave blank if all signal labels are 0.",
        "Example": "wait, no; actually; I don't know; flip it over; then; [Pause].",
    },
    {
        "Field": "Counterexample_or_Exclusion",
        "Short label": "Reject",
        "Required": "Optional",
        "Allowed values": "short text",
        "How to fill": "Evidence box 2: rejected cue. Record misleading cue-like language that you deliberately excluded from a signal label.",
        "Example": "looks like = visual comparison, not affective; no = answer choice label, not metacognitive reset.",
    },
    {
        "Field": "Codebook_Issue",
        "Short label": "Issue",
        "Required": "Optional",
        "Allowed values": "short text",
        "How to fill": "Evidence box 3: rule question. Describe an ambiguity in the rule, definition, or vocabulary that the research team should discuss.",
        "Example": "Need rule for when wait is M_t vs R_t; boundary valid but no signal definition fits.",
    },
    {
        "Field": "Notes",
        "Short label": "Notes",
        "Required": "Optional",
        "Allowed values": "free text",
        "How to fill": "Use for any extra explanation that does not fit the other fields.",
        "Example": "Boundary valid but signal type is hard to separate.",
    },
]


def annotation_short_labels_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Short label": row.get("Short label", ""),
                "Export column": row["Field"],
                "Required": row["Required"],
                "Meaning": row["How to fill"],
            }
            for row in ANNOTATION_CODEBOOK_ROWS
        ]
    )


def init_db() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS calibration_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                operation TEXT NOT NULL,
                input_path TEXT,
                output_path TEXT,
                parameters_json TEXT,
                status TEXT NOT NULL,
                message TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS current_artifacts (
                step TEXT PRIMARY KEY,
                updated_at TEXT NOT NULL,
                artifact_path TEXT,
                parameters_json TEXT,
                message TEXT
            )
            """
        )
        conn.commit()


def log_event(
    operation: str,
    input_path: str | Path | None,
    output_path: str | Path | None,
    parameters: dict,
    status: str,
    message: str,
) -> None:
    init_db()
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO calibration_events
            (timestamp, operation, input_path, output_path, parameters_json, status, message)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                operation,
                str(input_path) if input_path else "",
                str(output_path) if output_path else "",
                json.dumps(parameters, ensure_ascii=False, default=str),
                status,
                message,
            ),
        )
        conn.commit()


def replace_current_table(table_name: str, frame: pd.DataFrame) -> None:
    init_db()
    safe_name = re.sub(r"[^A-Za-z0-9_]+", "_", table_name).strip("_")
    if not safe_name:
        raise ValueError("SQLite table name cannot be empty.")
    with sqlite3.connect(DB_PATH) as conn:
        frame.copy().to_sql(safe_name, conn, if_exists="replace", index=False)


def save_current_artifact(
    step: str,
    artifact_path: str | Path | None,
    parameters: dict,
    message: str,
) -> None:
    init_db()
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO current_artifacts (step, updated_at, artifact_path, parameters_json, message)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(step) DO UPDATE SET
                updated_at=excluded.updated_at,
                artifact_path=excluded.artifact_path,
                parameters_json=excluded.parameters_json,
                message=excluded.message
            """,
            (
                step,
                datetime.now().isoformat(timespec="seconds"),
                str(artifact_path) if artifact_path else "",
                json.dumps(parameters, ensure_ascii=False, default=str),
                message,
            ),
        )
        conn.commit()


def read_current_artifacts() -> pd.DataFrame:
    init_db()
    with sqlite3.connect(DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT step, updated_at, artifact_path, parameters_json, message
            FROM current_artifacts
            ORDER BY step
            """,
            conn,
        )


def read_current_table(table_name: str) -> pd.DataFrame:
    init_db()
    safe_name = re.sub(r"[^A-Za-z0-9_]+", "_", table_name).strip("_")
    if not safe_name:
        return pd.DataFrame()
    try:
        with sqlite3.connect(DB_PATH) as conn:
            existing = pd.read_sql_query(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                conn,
                params=(safe_name,),
            )
            if existing.empty:
                return pd.DataFrame()
            return pd.read_sql_query(f'SELECT * FROM "{safe_name}"', conn)
    except Exception:
        return pd.DataFrame()


def current_artifact_path(step: str) -> str | None:
    artifacts = read_current_artifacts()
    if artifacts.empty:
        return None
    rows = artifacts[artifacts["step"] == step]
    if rows.empty:
        return None
    path = str(rows.iloc[0].get("artifact_path", "") or "")
    return path if path and Path(path).exists() else None


def refresh_sqlite_state(limit: int = 100) -> tuple[pd.DataFrame, pd.DataFrame]:
    return read_history(limit), read_current_artifacts()


def read_history(limit: int = 100) -> pd.DataFrame:
    init_db()
    with sqlite3.connect(DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT id, timestamp, operation, status, input_path, output_path, parameters_json, message
            FROM calibration_events
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(int(limit),),
        )


def step_header(title: str) -> gr.HTML:
    tooltip = STEP_HELP[title].replace('"', "&quot;")
    return gr.HTML(
        f'<div class="step-title">{title}<span class="help-dot" data-tooltip="{tooltip}">?</span></div>'
        f'<div class="step-note">{STEP_HELP[title]}</div>'
    )


STEP0_FIELD_HELP = {
    "Semantic continuity model": (
        "Formula role: S_t^sem. Choose how DAPS estimates local semantic continuity around a candidate boundary. "
        "lexical uses word overlap/string similarity and is recommended for calibration/debugging. "
        "all-mpnet-base-v2 uses sentence-embedding cosine similarity and is better for later comparison after human labels exist."
    ),
    "Clean encoding artifacts / mojibake": (
        "Preprocessing switch. Removes common encoding artifacts, corrupted characters, event leakage, and stable speaker-label text "
        "before segmentation. Recommended: on. This is rule-based cleanup, not LLM rewriting."
    ),
    "Context width": (
        "Formula symbol: w. Number of tokens on each side of gap g_t used to compute S_t^sem, D_t, C_t, M_t, A_t, and R_t. "
        "Suggested: w = 12."
    ),
    "Local radius": (
        "Formula symbol: r. Neighborhood size used for adaptive local valley detection in G'_t. "
        "Suggested: r = 6."
    ),
    "Sensitivity": (
        "Formula symbol: tau. Controls how deep a local semantic-gravity valley must be before becoming a candidate boundary. "
        "Higher tau usually means fewer segments. Suggested: tau = 0.55."
    ),
    "Minimum segment tokens": (
        "Formula symbol: L_min. Minimum allowed segment length during constrained decoding. "
        "Increasing it reduces oversegmentation. Suggested: L_min = 12."
    ),
    "Maximum segment tokens": (
        "Formula symbol: L_max. Preferred upper bound on segment length; longer spans trigger an additional legal split. "
        "Suggested: L_max = 60."
    ),
    "Maximum segments per record": (
        "Formula symbol: K_max. Record-level cap for extreme oversegmentation. "
        "Suggested: K_max = 80; set to 0 to disable."
    ),
    "NMS radius": (
        "Formula symbol: rho. Non-maximum suppression radius; nearby candidate boundaries compete so one local shift is not split multiple times. "
        "Suggested: rho = 6."
    ),
    "Task-density vocabulary": (
        "Formula symbol: V_task. Terms used to estimate task-word density around a candidate gap. "
        "Use domain/task objects and actions; keep broad filler words out."
    ),
    "Cognitive transition vocabulary": (
        "Formula symbol: V_C for C_t. Cues for shifts in object, action, strategy, comparison, or reasoning state. "
        "Suggested: prefer task-relevant action and relation terms over generic words."
    ),
    "Metacognitive reset vocabulary": (
        "Formula symbol: V_M for M_t. Cues for monitoring, correction, uncertainty, realization, or restart. "
        "Suggested: use phrases such as wait, actually, I don't know, I think, maybe with caution."
    ),
    "Affective friction vocabulary": (
        "Formula symbol: V_A for A_t. Cues for difficulty, frustration, confusion, effort, or affective resistance. "
        "Suggested: avoid ambiguous words unless human calibration supports them."
    ),
    "Rhetorical/structural break vocabulary": (
        "Formula symbol: V_R for R_t. Cues for discourse organization, sequencing, speaker/event shifts, and structural transitions. "
        "Suggested: then, next, okay, so, but, pause/event markers."
    ),
}


def help_label(name: str, symbol: str | None = None) -> gr.HTML:
    help_text = html.escape(STEP0_FIELD_HELP[name], quote=True)
    symbol_html = f'<span class="parameter-symbol">({html.escape(symbol)})</span>' if symbol else ""
    return gr.HTML(
        f'<div class="field-label">{html.escape(name)} {symbol_html}'
        f'<span class="help-dot" data-tooltip="{help_text}">?</span></div>'
    )


def step0_slider(
    name: str,
    symbol: str,
    minimum: float,
    maximum: float,
    value: float,
    step: float,
) -> gr.Slider:
    with gr.Column(elem_classes=["parameter-field"]):
        help_label(name, symbol)
        return gr.Slider(minimum, maximum, value=value, step=step, show_label=False)


def _split_columns(text: str | None) -> list[str] | None:
    if not text or not str(text).strip():
        return None
    return [col.strip() for col in str(text).split(",") if col.strip()]


def run_initial_segmentation(
    raw_file: str | None,
    sheet_name: str,
    id_column: str,
    item_column: str,
    text_columns: str,
    embedding_model: str,
    clean_text: bool,
    context_width: int,
    local_radius: int,
    sensitivity: float,
    min_segment_tokens: int,
    max_segment_tokens: int,
    max_segments_per_record: int,
    nms_radius: int,
    task_vocab: str,
    cognitive_vocab: str,
    metacognitive_vocab: str,
    affective_vocab: str,
    structural_vocab: str,
) -> tuple[str | None, pd.DataFrame, pd.DataFrame, str]:
    path = _path_from_upload(raw_file)
    if path is None:
        return None, pd.DataFrame(), pd.DataFrame(), "Upload a raw CSV/Excel transcript file first."

    try:
        sheet = sheet_name.strip() or None
        df = read_table(path, sheet)
        original_columns = list(df.columns)
        artifact_count = 0
        if clean_text:
            df, artifact_count = clean_dataframe_text(df)

        selected_id = id_column.strip() or infer_id_column(df, None)
        if selected_id and selected_id not in df.columns:
            raise ValueError(f"ID column not found: {selected_id}")

        selected_item = item_column.strip()
        if selected_item and selected_item not in df.columns:
            raise ValueError(f"Item/task column not found: {selected_item}")

        requested_text_cols = ",".join(_split_columns(text_columns) or []) or None
        selected_text_columns = infer_text_columns(df, requested_text_cols, selected_id)
        vocabulary = SignalVocabulary.from_texts(
            task=task_vocab,
            cognitive=cognitive_vocab,
            metacognitive=metacognitive_vocab,
            affective=affective_vocab,
            structural=structural_vocab,
        )

        segmenter = DAPSSegmenter(
            DAPSConfig(
                context_width=int(context_width),
                local_radius=int(local_radius),
                sensitivity=float(sensitivity),
                min_segment_tokens=int(min_segment_tokens),
                max_segment_tokens=int(max_segment_tokens),
                max_segments_per_record=int(max_segments_per_record),
                nms_radius=int(nms_radius),
                cognitive_drop_floor=0.35,
                cognitive_drop_ceiling=0.80,
                cognitive_semantic_weight=0.45,
                cognitive_cue_shift_weight=0.55,
                cognitive_semantic_gate_floor=0.25,
                legality_mode="hybrid",
                spacy_model="en_core_web_sm",
                semi_legal_penalty=0.18,
                sentence_boundary_boost=0.08,
            ),
            SimilarityModel(embedding_model.strip() or "lexical"),
            vocabulary,
        )

        metadata_columns = [
            col for col in [selected_id, selected_item, "Response"] if col and col in df.columns
        ]
        all_segments = []
        all_boundaries = []
        for row_index, row in df.iterrows():
            id_value = row[selected_id] if selected_id else row_index + 1
            item_value = row[selected_item] if selected_item else ""
            record_id = f"{id_value}__{item_value}" if selected_item else id_value
            metadata = {
                f"Source_{col}": row[col]
                for col in metadata_columns
                if col in row.index
            }
            for source_column in selected_text_columns:
                text = normalize_transcript(row[source_column], clean_text=clean_text)
                if not text:
                    continue
                segments, boundaries = segmenter.segment(text)
                for segment in segments:
                    all_segments.append(
                        {
                            "Record_ID": record_id,
                            "Source_Row": row_index + 2,
                            "Source_Column": source_column,
                            **metadata,
                            **segment,
                        }
                    )
                for boundary in boundaries:
                    all_boundaries.append(
                        {
                            "Record_ID": record_id,
                            "Source_Row": row_index + 2,
                            "Source_Column": source_column,
                            **metadata,
                            **boundary,
                        }
                    )

        segments_df = pd.DataFrame(all_segments)
        boundaries_df = pd.DataFrame(all_boundaries)
        if segments_df.empty:
            raise ValueError("No segments were generated. Check the transcript text column.")

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        out = OUTPUT_DIR / f"{path.stem}_initial_daps_segments.xlsx"
        metadata_df = pd.DataFrame(
            [
                {"Field": "Input file", "Value": str(path)},
                {"Field": "Original columns", "Value": ", ".join(map(str, original_columns))},
                {"Field": "ID column", "Value": selected_id or ""},
                {"Field": "Item/task column", "Value": selected_item or ""},
                {"Field": "Text columns", "Value": ", ".join(selected_text_columns)},
                {"Field": "Clean text", "Value": clean_text},
                {"Field": "Encoding/artifact count cleaned", "Value": artifact_count},
                {"Field": "Task vocabulary size", "Value": len(vocabulary.task)},
                {"Field": "C_t vocabulary size", "Value": len(vocabulary.cognitive)},
                {"Field": "M_t vocabulary size", "Value": len(vocabulary.metacognitive)},
                {"Field": "A_t vocabulary size", "Value": len(vocabulary.affective)},
                {"Field": "R_t vocabulary size", "Value": len(vocabulary.structural)},
                {"Field": "Segments", "Value": len(segments_df)},
                {"Field": "Boundaries", "Value": len(boundaries_df)},
            ]
        )
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            segments_df.to_excel(writer, sheet_name="segments", index=False)
            boundaries_df.to_excel(writer, sheet_name="boundaries", index=False)
            df.to_excel(writer, sheet_name="cleaned_input", index=False)
            metadata_df.to_excel(writer, sheet_name="run_metadata", index=False)

        status = (
            f"Segmented {len(df)} source rows into {len(segments_df)} segments and "
            f"{len(boundaries_df)} boundary evidence rows. Cleaned {artifact_count} encoding artifact(s). "
            f"Vocabulary sizes: C={len(vocabulary.cognitive)}, M={len(vocabulary.metacognitive)}, "
            f"A={len(vocabulary.affective)}, R={len(vocabulary.structural)}, Task={len(vocabulary.task)}."
        )
        parameters = {
            "id_column": selected_id,
            "item_column": selected_item,
            "text_columns": selected_text_columns,
            "embedding_model": embedding_model,
            "clean_text": clean_text,
            "context_width": context_width,
            "local_radius": local_radius,
            "sensitivity": sensitivity,
            "min_segment_tokens": min_segment_tokens,
            "max_segment_tokens": max_segment_tokens,
            "max_segments_per_record": max_segments_per_record,
            "nms_radius": nms_radius,
            "task_vocab_size": len(vocabulary.task),
            "cognitive_vocab_size": len(vocabulary.cognitive),
            "metacognitive_vocab_size": len(vocabulary.metacognitive),
            "affective_vocab_size": len(vocabulary.affective),
            "structural_vocab_size": len(vocabulary.structural),
        }
        replace_current_table("step0_segments", segments_df)
        replace_current_table("step0_boundaries", boundaries_df)
        replace_current_table("step0_cleaned_input", df)
        replace_current_table("step0_run_metadata", metadata_df)
        save_current_artifact("step0_initial_segmentation", out, parameters, status)
        log_event(
            "initial_segmentation",
            path,
            out,
            parameters,
            "success",
            status,
        )
        return str(out), segments_df.head(30), boundaries_df.head(30), status
    except Exception as exc:
        message = f"Initial segmentation failed: {exc}"
        log_event("initial_segmentation", path, "", {"input": str(path)}, "error", message)
        return None, pd.DataFrame(), pd.DataFrame(), message


def run_initial_segmentation_and_inspect(
    raw_file: str | None,
    sheet_name: str,
    id_column: str,
    item_column: str,
    text_columns: str,
    embedding_model: str,
    clean_text: bool,
    context_width: int,
    local_radius: int,
    sensitivity: float,
    min_segment_tokens: int,
    max_segment_tokens: int,
    max_segments_per_record: int,
    nms_radius: int,
    task_vocab: str,
    cognitive_vocab: str,
    metacognitive_vocab: str,
    affective_vocab: str,
    structural_vocab: str,
) -> tuple[
    str | None,
    pd.DataFrame,
    pd.DataFrame,
    str,
    str | None,
    str | None,
    pd.DataFrame,
    pd.DataFrame,
    str,
    str,
    str,
]:
    output_path, segments_preview, boundaries_preview, status = run_initial_segmentation(
        raw_file,
        sheet_name,
        id_column,
        item_column,
        text_columns,
        embedding_model,
        clean_text,
        context_width,
        local_radius,
        sensitivity,
        min_segment_tokens,
        max_segment_tokens,
        max_segments_per_record,
        nms_radius,
        task_vocab,
        cognitive_vocab,
        metacognitive_vocab,
        affective_vocab,
        structural_vocab,
    )
    if not output_path:
        return (
            output_path,
            segments_preview,
            boundaries_preview,
            status,
            None,
            None,
            pd.DataFrame(),
            pd.DataFrame(),
            "",
            "",
            "Step 1 inspection skipped because Step 0 failed.",
        )

    (
        inspection_summary,
        signal_summary,
        inspection_summary_html,
        signal_summary_html,
        inspection_status,
    ) = inspect_segmentation_workbook_for_ui(output_path)
    combined_status = f"{status}\n\nStep 1 auto-inspection completed. Use the generated workbook for Step 2."
    return (
        output_path,
        segments_preview,
        boundaries_preview,
        combined_status,
        output_path,
        output_path,
        inspection_summary,
        signal_summary,
        inspection_summary_html,
        signal_summary_html,
        inspection_status,
    )


def load_latest_sqlite_state() -> tuple[
    str | None,
    pd.DataFrame,
    pd.DataFrame,
    str,
    str | None,
    str | None,
    pd.DataFrame,
    pd.DataFrame,
    str,
    str,
    str,
    pd.DataFrame,
    pd.DataFrame,
]:
    segments = read_current_table("step0_segments")
    boundaries = read_current_table("step0_boundaries")
    summary = read_current_table("step1_prevalidation_summary")
    signals = read_current_table("step1_signal_saturation_summary")
    history, artifacts = refresh_sqlite_state(100)

    segmented_path = current_artifact_path("step0_initial_segmentation")
    step0_message = "No Step 0 segmentation found in SQLite yet."
    if not segments.empty:
        step0_message = (
            f"Loaded latest Step 0 data from SQLite: {len(segments)} segments and "
            f"{len(boundaries)} boundary evidence rows. No re-segmentation was run."
        )
        if segmented_path:
            step0_message += f"\nCurrent workbook: {segmented_path}"
        else:
            step0_message += "\nWorkbook path was not found, but SQLite tables are available."

    inspection_message = "No Step 1 inspection summary found in SQLite yet."
    if not summary.empty or not signals.empty:
        inspection_message = "Loaded latest Step 1 inspection from SQLite. No segmentation was run."

    return (
        segmented_path,
        segments.head(30),
        boundaries.head(30),
        step0_message,
        segmented_path,
        segmented_path,
        summary,
        signals,
        _status_table_html(summary, "Pre-validation status"),
        _status_table_html(signals, "Signal saturation status"),
        inspection_message,
        history,
        artifacts,
    )


def _path_from_upload(uploaded_file: str | Path | None) -> Path | None:
    if uploaded_file is None:
        return None
    if isinstance(uploaded_file, dict):
        for key in ("path", "name", "orig_name"):
            value = uploaded_file.get(key)
            if value:
                return Path(value)
    if hasattr(uploaded_file, "path"):
        return Path(uploaded_file.path)
    if hasattr(uploaded_file, "name"):
        return Path(uploaded_file.name)
    return Path(uploaded_file)


def _bool_series(series: pd.Series) -> pd.Series:
    if series.dtype == bool:
        return series.fillna(False)
    return series.fillna(False).astype(str).str.lower().isin({"true", "1", "yes", "y"})


def _to_float(value: object) -> float | None:
    try:
        if pd.isna(value):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _status_for_prevalidation_metric(metric: str, value: object) -> tuple[str, str]:
    numeric = _to_float(value)
    if metric in {"Segments", "Boundaries", "Records", "Selected boundaries"}:
        return "INFO", "Reported for traceability; no universal pass/fail threshold."
    if numeric is None:
        return "WARNING", "Missing or non-numeric value; inspect workbook columns."
    if metric == "Median segment tokens":
        return (
            ("PASS", "Median length is in the suggested 12-25 token range.")
            if 12 <= numeric <= 25
            else ("WARNING", "Median length is outside the suggested 12-25 token range.")
        )
    if metric == "Mean segment tokens":
        return (
            ("PASS", "Mean length is in the suggested 15-30 token range.")
            if 15 <= numeric <= 30
            else ("WARNING", "Mean length is outside the suggested 15-30 token range.")
        )
    if metric == "Short segments < 8 tokens":
        return "INFO", "Use Short segment rate for pass/fail; this is the raw count."
    if metric == "Short segment rate":
        return (
            ("PASS", "Short segment rate is within the suggested maximum.")
            if numeric <= 0.25
            else ("WARNING", "Short segment rate is high; segmentation may be too fine.")
        )
    if metric == "Long segments > 60 tokens":
        return (
            ("PASS", "No excessive long segments detected.")
            if numeric == 0
            else ("WARNING", "Long segments remain; consider lowering L_max or adding forced splitting.")
        )
    if metric == "Max tokens":
        return (
            ("PASS", "Maximum segment length is within L_max.")
            if numeric <= 60
            else ("WARNING", "Maximum segment length exceeds the default L_max=60 criterion.")
        )
    if metric == "Max segments per record":
        return (
            ("PASS", "Maximum segments per record is within K_max.")
            if numeric <= 80
            else ("WARNING", "Some records have many segments; check oversegmentation.")
        )
    if metric in {
        "Event token leakage",
        "Bracket event leakage",
        "Speaker label leakage",
        "Punctuation-leading segments",
    }:
        return (
            ("PASS", "No leakage detected.")
            if numeric == 0
            else ("FAIL", "Cleanup or post-processing should remove these cases before annotation.")
        )
    if metric == "Selected illegal boundaries":
        return (
            ("PASS", "No selected boundaries violate the legality filter.")
            if numeric == 0
            else ("FAIL", "Illegal selected boundaries remain; review legality rules or forced max-length splitting.")
        )
    if metric == "Selected semi-legal boundary rate":
        return (
            ("PASS", "Few selected boundaries are only semi-legal.")
            if numeric <= 0.15
            else ("WARNING", "Many selected boundaries are semi-legal; increase lambda_S or inspect parser/rule settings.")
        )
    return "INFO", "Pairwise criterion evaluated after selected/non-selected rows are compared."


def _apply_prevalidation_status(summary: pd.DataFrame) -> pd.DataFrame:
    summary = summary.copy()
    statuses = []
    notes = []
    for _, row in summary.iterrows():
        status, note = _status_for_prevalidation_metric(str(row["Metric"]), row["Value"])
        statuses.append(status)
        notes.append(note)

    def set_pair_status(selected_metric: str, nonselected_metric: str, selected_should_be_lower: bool) -> None:
        selected_idx = summary.index[summary["Metric"] == selected_metric]
        nonselected_idx = summary.index[summary["Metric"] == nonselected_metric]
        if selected_idx.empty or nonselected_idx.empty:
            return
        selected_value = _to_float(summary.loc[selected_idx[0], "Value"])
        nonselected_value = _to_float(summary.loc[nonselected_idx[0], "Value"])
        if selected_value is None or nonselected_value is None:
            pair_status = "WARNING"
            pair_note = "Cannot compare selected and non-selected values."
        else:
            passes = selected_value < nonselected_value if selected_should_be_lower else selected_value > nonselected_value
            pair_status = "PASS" if passes else "WARNING"
            direction = "lower" if selected_should_be_lower else "higher"
            pair_note = (
                f"Selected boundaries are {direction} than non-selected boundaries as expected."
                if passes
                else f"Selected boundaries are not {direction} than non-selected boundaries; recalibrate signals."
            )
        for idx in (selected_idx[0], nonselected_idx[0]):
            statuses[idx] = pair_status
            notes[idx] = pair_note

    set_pair_status(
        "Mean semantic gravity: selected",
        "Mean semantic gravity: non-selected",
        selected_should_be_lower=True,
    )
    set_pair_status(
        "Mean transition pressure: selected",
        "Mean transition pressure: non-selected",
        selected_should_be_lower=False,
    )

    summary["Status"] = statuses
    summary["Status_Note"] = notes
    return summary


def _apply_signal_status(signals: pd.DataFrame) -> pd.DataFrame:
    signals = signals.copy()
    statuses = []
    notes = []
    for _, row in signals.iterrows():
        median = _to_float(row.get("Median"))
        pct_high = _to_float(row.get("Pct > .80")) or 0.0
        saturated = bool(row.get("Saturation warning", False))
        if median is None:
            statuses.append("WARNING")
            notes.append("Missing signal values.")
        elif saturated or median > 0.75 or pct_high > 0.50:
            statuses.append("WARNING")
            notes.append("Signal may be saturated; reduce cue breadth or weight before vocabulary training.")
        else:
            statuses.append("PASS")
            notes.append("Signal distribution has usable separation for calibration.")
    signals["Status"] = statuses
    signals["Status_Note"] = notes
    return signals


def _status_table_html(frame: pd.DataFrame, title: str) -> str:
    if frame.empty:
        return ""
    display_columns = [
        col
        for col in frame.columns
        if col
        in {
            "Metric",
            "Signal",
            "Value",
            "Mean",
            "Median",
            "P90",
            "Pct > .80",
            "Suggested criterion",
            "Saturation warning",
            "Status",
            "Status_Note",
        }
    ]
    header = "".join(f"<th>{html.escape(str(column))}</th>" for column in display_columns)
    rows = []
    for _, row in frame.iterrows():
        status = str(row.get("Status", "INFO")).lower()
        cells = []
        for column in display_columns:
            value = row.get(column, "")
            text = "" if pd.isna(value) else str(value)
            if column == "Status":
                text = f'<span class="status-badge">{html.escape(text)}</span>'
            elif column in {"Metric", "Signal"}:
                help_text = METRIC_HELP.get(str(value), SIGNAL_HELP.get(str(value), "No additional description is available."))
                text = (
                    f"{html.escape(text)}"
                    f'<span class="metric-help-dot" data-tooltip="{html.escape(help_text, quote=True)}">?</span>'
                )
            else:
                text = html.escape(text)
            cells.append(f"<td>{text}</td>")
        rows.append(f'<tr class="status-{html.escape(status)}">{"".join(cells)}</tr>')
    return f"<h4>{html.escape(title)}</h4><table class=\"status-table\"><thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody></table>"


def _parse_annotators(text: str) -> list[str]:
    annotators = [item.strip() for item in text.replace(";", ",").split(",") if item.strip()]
    return list(dict.fromkeys(annotators)) or ["Annotator_1"]


def _dominant_signal(row: pd.Series) -> str:
    scores = {signal: float(row.get(column, 0.0) or 0.0) for signal, column in SIGNAL_COLUMNS.items()}
    best_signal, best_score = max(scores.items(), key=lambda item: item[1])
    return best_signal if best_score >= 0.20 else "low_signal"


def _find_segment_context(segments: pd.DataFrame, boundary: pd.Series) -> tuple[str, str, str]:
    mask = (
        segments["Record_ID"].eq(boundary["Record_ID"])
        & segments["Source_Column"].eq(boundary["Source_Column"])
    )
    source_segments = segments.loc[mask].sort_values("Segment_Index")
    gap = boundary["gap"]
    left = source_segments[source_segments["Token_End"].eq(gap)]
    right = source_segments[source_segments["Token_Start"].eq(gap)]

    if not left.empty:
        left_row = left.iloc[-1]
        previous = source_segments[source_segments["Segment_Index"].eq(left_row["Segment_Index"] - 1)]
        previous_text = "" if previous.empty else str(previous.iloc[0]["Segment_Text"])
        left_text = str(left_row["Segment_Text"])
    else:
        previous_text = ""
        before = source_segments[source_segments["Token_End"] < gap]
        left_text = "" if before.empty else str(before.iloc[-1]["Segment_Text"])

    if not right.empty:
        right_text = str(right.iloc[0]["Segment_Text"])
    else:
        after = source_segments[source_segments["Token_Start"] > gap]
        right_text = "" if after.empty else str(after.iloc[0]["Segment_Text"])

    return previous_text, left_text, right_text


def load_segmentation_workbook(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    segments = pd.read_excel(path, sheet_name="segments")
    boundaries = pd.read_excel(path, sheet_name="boundaries")
    required_segments = {"Record_ID", "Source_Column", "Segment_Index", "Token_Start", "Token_End", "Segment_Text"}
    required_boundaries = {"Record_ID", "Source_Column", "gap", "candidate_boundary", "selected_boundary"}
    missing_segments = required_segments - set(segments.columns)
    missing_boundaries = required_boundaries - set(boundaries.columns)
    if missing_segments:
        raise ValueError(f"segments sheet missing columns: {', '.join(sorted(missing_segments))}")
    if missing_boundaries:
        raise ValueError(f"boundaries sheet missing columns: {', '.join(sorted(missing_boundaries))}")
    return segments, boundaries


def inspect_segmentation_workbook(segmentation_workbook: str | None) -> tuple[pd.DataFrame, pd.DataFrame, str]:
    path = _path_from_upload(segmentation_workbook)
    if path is None:
        return pd.DataFrame(), pd.DataFrame(), "Upload a DAPS segmentation workbook first."
    try:
        segments, boundaries = load_segmentation_workbook(path)
        texts = segments["Segment_Text"].fillna("").astype(str)
        selected = _bool_series(boundaries["selected_boundary"])
        non_selected = ~selected

        summary_rows = [
            {"Metric": "Segments", "Value": len(segments), "Suggested criterion": "Documented"},
            {"Metric": "Boundaries", "Value": len(boundaries), "Suggested criterion": "Documented"},
            {"Metric": "Records", "Value": segments["Record_ID"].nunique(), "Suggested criterion": "Documented"},
            {"Metric": "Median segment tokens", "Value": round(float(segments["Token_Count"].median()), 2), "Suggested criterion": "12-25"},
            {"Metric": "Mean segment tokens", "Value": round(float(segments["Token_Count"].mean()), 2), "Suggested criterion": "15-30"},
            {"Metric": "Short segments < 8 tokens", "Value": int((segments["Token_Count"] < 8).sum()), "Suggested criterion": "< 20-25%"},
            {"Metric": "Short segment rate", "Value": round(float((segments["Token_Count"] < 8).mean()), 3), "Suggested criterion": "< .25"},
            {"Metric": "Long segments > 60 tokens", "Value": int((segments["Token_Count"] > 60).sum()), "Suggested criterion": "0 or near 0"},
            {"Metric": "Max tokens", "Value": int(segments["Token_Count"].max()), "Suggested criterion": "<= L_max"},
            {"Metric": "Max segments per record", "Value": int(segments.groupby("Record_ID").size().max()), "Suggested criterion": "<= K_max"},
            {"Metric": "Event token leakage", "Value": int(texts.str.contains("<EVENT>", regex=False).sum()), "Suggested criterion": "0"},
            {"Metric": "Bracket event leakage", "Value": int(texts.str.contains(r"\[(?:Pause|Drawing|End of Audio|Unintelligible)", regex=True, case=False).sum()), "Suggested criterion": "0"},
            {"Metric": "Speaker label leakage", "Value": int(texts.str.contains(SPEAKER_LABEL_RE).sum()), "Suggested criterion": "0"},
            {"Metric": "Punctuation-leading segments", "Value": int(texts.str.match(r"^\s*[,.;:!?]", na=False).sum()), "Suggested criterion": "0"},
            {"Metric": "Selected boundaries", "Value": int(selected.sum()), "Suggested criterion": "Documented"},
        ]

        if "boundary_legality" in boundaries.columns:
            legality = boundaries["boundary_legality"].fillna("").astype(str)
            selected_count = max(1, int(selected.sum()))
            summary_rows.extend(
                [
                    {
                        "Metric": "Selected illegal boundaries",
                        "Value": int(((legality == "illegal") & selected).sum()),
                        "Suggested criterion": "0",
                    },
                    {
                        "Metric": "Selected semi-legal boundary rate",
                        "Value": round(float(((legality == "semi_legal") & selected).sum()) / selected_count, 3),
                        "Suggested criterion": "<= .15",
                    },
                ]
            )

        if {"semantic_gravity", "transition_pressure"}.issubset(boundaries.columns):
            summary_rows.extend(
                [
                    {
                        "Metric": "Mean semantic gravity: selected",
                        "Value": round(float(boundaries.loc[selected, "semantic_gravity"].mean()), 4),
                        "Suggested criterion": "Lower than non-selected",
                    },
                    {
                        "Metric": "Mean semantic gravity: non-selected",
                        "Value": round(float(boundaries.loc[non_selected, "semantic_gravity"].mean()), 4),
                        "Suggested criterion": "Higher than selected",
                    },
                    {
                        "Metric": "Mean transition pressure: selected",
                        "Value": round(float(boundaries.loc[selected, "transition_pressure"].mean()), 4),
                        "Suggested criterion": "Higher than non-selected",
                    },
                    {
                        "Metric": "Mean transition pressure: non-selected",
                        "Value": round(float(boundaries.loc[non_selected, "transition_pressure"].mean()), 4),
                        "Suggested criterion": "Lower than selected",
                    },
                ]
            )

        signal_rows = []
        for signal, column in SIGNAL_COLUMNS.items():
            if column in boundaries.columns:
                values = boundaries[column].dropna()
                signal_rows.append(
                    {
                        "Signal": signal,
                        "Mean": round(float(values.mean()), 3),
                        "Median": round(float(values.median()), 3),
                        "P90": round(float(values.quantile(0.90)), 3),
                        "Pct > .80": round(float((values > 0.80).mean()), 3),
                        "Saturation warning": bool(values.median() > 0.75),
                    }
                )

        summary = _apply_prevalidation_status(pd.DataFrame(summary_rows))
        signals = _apply_signal_status(pd.DataFrame(signal_rows))
        status = f"Inspected {path.name}. Pre-validation summary created and logged to SQLite."
        parameters = {"workbook": str(path)}
        replace_current_table("step1_prevalidation_summary", summary)
        replace_current_table("step1_signal_saturation_summary", signals)
        save_current_artifact("step1_inspection", path, parameters, status)
        log_event(
            "inspect_workbook",
            path,
            "",
            parameters,
            "success",
            status,
        )
        return summary, signals, status
    except Exception as exc:
        message = f"Inspection failed: {exc}"
        log_event("inspect_workbook", path, "", {"workbook": str(path)}, "error", message)
        return pd.DataFrame(), pd.DataFrame(), message


def inspect_segmentation_workbook_for_ui(segmentation_workbook: str | None) -> tuple[pd.DataFrame, pd.DataFrame, str, str, str]:
    summary, signals, status = inspect_segmentation_workbook(segmentation_workbook)
    return (
        summary,
        signals,
        _status_table_html(summary, "Pre-validation status"),
        _status_table_html(signals, "Signal saturation status"),
        status,
    )


def build_boundary_items(boundaries: pd.DataFrame) -> pd.DataFrame:
    constrained = boundaries.get("constraint_reason", pd.Series("", index=boundaries.index)).fillna("").astype(str).str.strip().ne("")
    candidates = boundaries[
        _bool_series(boundaries["selected_boundary"])
        | (_bool_series(boundaries["candidate_boundary"]) & ~constrained)
    ].copy()
    candidates["Item_Source"] = np.where(_bool_series(candidates["selected_boundary"]), "selected", "candidate")
    candidates["Algorithm_Dominant_Signal"] = candidates.apply(_dominant_signal, axis=1)
    candidates["Sampling_Stratum"] = candidates["Item_Source"] + "_" + candidates["Algorithm_Dominant_Signal"]

    rows = []
    for index, row in candidates.reset_index(drop=True).iterrows():
        item_id = f"{row['Record_ID']}__row_{row.get('Source_Row', '')}__{row['Source_Column']}__gap_{int(row['gap'])}"
        rows.append(
            {
                "Annotation_Item_ID": item_id,
                "Record_ID": row["Record_ID"],
                "Source_Row": row.get("Source_Row", ""),
                "Source_Column": row["Source_Column"],
                "Boundary_Gap": int(row["gap"]),
                "Item_Source": row["Item_Source"],
                "Sampling_Stratum": row["Sampling_Stratum"],
                "Algorithm_Dominant_Signal": row["Algorithm_Dominant_Signal"],
                "Algorithm_Selected_Boundary": bool(row.get("selected_boundary", False)),
                "Boundary_Legality": row.get("boundary_legality", ""),
                "Legality_Reason": row.get("legality_reason", ""),
                "Rule_Reason": row.get("rule_constraint_reason", ""),
                "Spacy_Reason": row.get("spacy_legality_reason", ""),
                "Spacy_Left_POS": row.get("spacy_left_pos", ""),
                "Spacy_Right_POS": row.get("spacy_right_pos", ""),
                "Spacy_Left_DEP": row.get("spacy_left_dep", ""),
                "Spacy_Right_DEP": row.get("spacy_right_dep", ""),
                "Boundary_Constraint_Reason": row.get("constraint_reason", ""),
                "Boundary_Marker_Explanation": "[[CANDIDATE_BOUNDARY]] marks the candidate break; it is not transcript text.",
                "Boundary_Context": str(row.get("context", "")).replace("<BOUNDARY>", "[[CANDIDATE_BOUNDARY]]"),
                "Alg_C_t": row.get("cognitive_transition", np.nan),
                "Alg_M_t": row.get("metacognitive_reset", np.nan),
                "Alg_A_t": row.get("affective_friction", np.nan),
                "Alg_R_t": row.get("structural_break", np.nan),
                "Semantic_Gravity": row.get("semantic_gravity", np.nan),
                "Boundary_Margin": row.get("boundary_margin", np.nan),
            }
        )
    return pd.DataFrame(rows)


def add_segment_context(items: pd.DataFrame, segments: pd.DataFrame) -> pd.DataFrame:
    enriched = items.copy()
    previous_values = []
    left_values = []
    right_values = []
    for _, row in enriched.iterrows():
        previous, left, right = _find_segment_context(
            segments,
            pd.Series(
                {
                    "Record_ID": row["Record_ID"],
                    "Source_Column": row["Source_Column"],
                    "gap": row["Boundary_Gap"],
                }
            ),
        )
        previous_values.append(previous)
        left_values.append(left)
        right_values.append(right)
    enriched["Previous_Segment"] = previous_values
    enriched["Left_Segment"] = left_values
    enriched["Right_Segment"] = right_values
    return enriched


def stratified_sample(items: pd.DataFrame, sample_size: int, seed: int) -> pd.DataFrame:
    if len(items) <= sample_size:
        return items.sample(frac=1.0, random_state=seed).reset_index(drop=True)

    strata = items.groupby("Sampling_Stratum", group_keys=False)
    per_stratum = max(1, sample_size // max(1, strata.ngroups))
    sampled = strata.apply(lambda group: group.sample(min(len(group), per_stratum), random_state=seed))
    if len(sampled) < sample_size:
        remaining = items.drop(sampled.index, errors="ignore")
        extra = remaining.sample(min(len(remaining), sample_size - len(sampled)), random_state=seed)
        sampled = pd.concat([sampled, extra], ignore_index=False)
    return sampled.sample(frac=1.0, random_state=seed).head(sample_size).reset_index(drop=True)


def _format_annotation_workbook(writer: pd.ExcelWriter, template: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    worksheet = writer.sheets["annotation_items"]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    max_row = len(template) + 1
    columns = {name: index + 1 for index, name in enumerate(template.columns)}

    for column_name, width in {
        "Annotation_ID": 34,
        "Annotation_Item_ID": 34,
        "Annotator_ID": 16,
        "Boundary_Strength_0_3": 20,
        "Human_Primary_Type": 20,
        "Cue_Span": 28,
        "Counterexample_or_Exclusion": 34,
        "Codebook_Issue": 34,
        "Notes": 34,
        "Previous_Segment": 46,
        "Left_Segment": 46,
        "Right_Segment": 46,
        "Boundary_Context": 60,
    }.items():
        if column_name in columns:
            worksheet.column_dimensions[get_column_letter(columns[column_name])].width = width

    binary_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    strength_validation = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
    primary_validation = DataValidation(
        type="list",
        formula1='"cognitive,metacognitive,affective,structural,mixed,low_signal,unclear"',
        allow_blank=True,
    )
    worksheet.add_data_validation(binary_validation)
    worksheet.add_data_validation(strength_validation)
    worksheet.add_data_validation(primary_validation)

    for column_name in ["Human_C_t", "Human_M_t", "Human_A_t", "Human_R_t"]:
        if column_name in columns and max_row >= 2:
            letter = get_column_letter(columns[column_name])
            binary_validation.add(f"{letter}2:{letter}{max_row}")
    if "Boundary_Strength_0_3" in columns and max_row >= 2:
        letter = get_column_letter(columns["Boundary_Strength_0_3"])
        strength_validation.add(f"{letter}2:{letter}{max_row}")
    if "Human_Primary_Type" in columns and max_row >= 2:
        letter = get_column_letter(columns["Human_Primary_Type"])
        primary_validation.add(f"{letter}2:{letter}{max_row}")

    for sheet_name in ["instructions", "codebook", "short_labels"]:
        if sheet_name not in writer.sheets:
            continue
        sheet = writer.sheets[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            width = 18 if header in {"Section", "Step", "Field", "Short label", "Required", "Allowed values"} else 78
            sheet.column_dimensions[column_cells[0].column_letter].width = width
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 42 if row[0].row > 1 else 24


def make_annotation_template(
    segmentation_workbook: str | None,
    calibration_round: str,
    sample_size: int,
    annotators: str,
    seed: int,
) -> tuple[str | None, pd.DataFrame, str]:
    path = _path_from_upload(segmentation_workbook)
    if path is None:
        return None, pd.DataFrame(), "Upload a DAPS segmentation workbook first."

    try:
        segments, boundaries = load_segmentation_workbook(path)
        items = build_boundary_items(boundaries)
        sampled = stratified_sample(items, int(sample_size), int(seed))
        sampled["Sampling_Stratum"] = sampled["Item_Source"] + "_" + sampled["Algorithm_Dominant_Signal"]
        sampled = add_segment_context(sampled, segments)
        annotator_ids = _parse_annotators(annotators)

        rows = []
        for annotator_id in annotator_ids:
            for _, item in sampled.iterrows():
                row = item.to_dict()
                row.update(
                    {
                        "Annotator_ID": annotator_id,
                        "Annotation_ID": f"{item['Annotation_Item_ID']}__{annotator_id}",
                        "Boundary_Strength_0_3": "",
                        "Human_C_t": "",
                        "Human_M_t": "",
                        "Human_A_t": "",
                        "Human_R_t": "",
                        "Human_Primary_Type": "",
                        "Cue_Span": "",
                        "Counterexample_or_Exclusion": "",
                        "Codebook_Issue": "",
                        "Notes": "",
                    }
                )
                rows.append(row)
        template = pd.DataFrame(rows)
        front_cols = [
            "Annotation_ID",
            "Annotation_Item_ID",
            "Annotator_ID",
            "Boundary_Strength_0_3",
            "Human_C_t",
            "Human_M_t",
            "Human_A_t",
            "Human_R_t",
            "Human_Primary_Type",
            "Cue_Span",
            "Counterexample_or_Exclusion",
            "Codebook_Issue",
            "Notes",
            "Sampling_Stratum",
        ]
        template = template[front_cols + [col for col in template.columns if col not in front_cols]]

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        round_slug = re.sub(r"[^A-Za-z0-9]+", "_", str(calibration_round).strip().lower()).strip("_") or "calibration"
        out = OUTPUT_DIR / f"{path.stem}_{round_slug}_annotation_template_{len(sampled)}x{len(annotator_ids)}.xlsx"
        instructions_df = pd.DataFrame(ANNOTATION_INSTRUCTIONS)
        codebook_df = pd.DataFrame(ANNOTATION_CODEBOOK_ROWS)
        short_labels_df = annotation_short_labels_df()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            template.to_excel(writer, sheet_name="annotation_items", index=False)
            instructions_df.to_excel(writer, sheet_name="instructions", index=False)
            codebook_df.to_excel(writer, sheet_name="codebook", index=False)
            short_labels_df.to_excel(writer, sheet_name="short_labels", index=False)
            _format_annotation_workbook(writer, template)

        status = (
            f"Generated {len(template)} annotation rows from {len(sampled)} unique items "
            f"for {len(annotator_ids)} annotator(s) in {calibration_round} mode."
        )
        parameters = {
            "calibration_round": calibration_round,
            "sample_size": int(sample_size),
            "annotators": annotator_ids,
            "seed": int(seed),
        }
        replace_current_table("step2_annotation_items", template)
        replace_current_table("step2_codebook", codebook_df)
        save_current_artifact("step2_annotation_template", out, parameters, status)
        log_event(
            "generate_annotation_template",
            path,
            out,
            parameters,
            "success",
            status,
        )
        return str(out), template.head(30), status
    except Exception as exc:
        message = f"Template generation failed: {exc}"
        log_event(
            "generate_annotation_template",
            path,
            "",
            {
                "calibration_round": calibration_round,
                "sample_size": sample_size,
                "annotators": annotators,
                "seed": seed,
            },
            "error",
            message,
        )
        return None, pd.DataFrame(), message


REVIEW_PRIMARY_TYPES = ["cognitive", "metacognitive", "affective", "structural", "mixed", "low_signal", "unclear"]
DEPRECATED_BOUNDARY_COLUMNS = ["Human_Boundary_Valid", "Confidence_1_3"]


def _single_boundary_variable_frame(frame: pd.DataFrame) -> pd.DataFrame:
    cleaned = frame.copy()
    if "Boundary_Strength_0_3" not in cleaned.columns and "Human_Boundary_Valid" in cleaned.columns:
        cleaned["Boundary_Strength_0_3"] = cleaned["Human_Boundary_Valid"].fillna("").astype(str).str.strip().map(
            {"0": "0", "1": "2"}
        )
    return cleaned.drop(columns=[col for col in DEPRECATED_BOUNDARY_COLUMNS if col in cleaned.columns])


def _review_dataframe() -> pd.DataFrame:
    frame = read_current_table("step2_annotation_items")
    if not frame.empty:
        return _single_boundary_variable_frame(frame)
    template_path = current_artifact_path("step2_annotation_template")
    if template_path:
        try:
            return _single_boundary_variable_frame(pd.read_excel(template_path, sheet_name="annotation_items"))
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()


def _review_choices(frame: pd.DataFrame) -> list[str]:
    if frame.empty or "Annotation_ID" not in frame.columns:
        return []
    choices = []
    total = len(frame)
    for index, row in frame.reset_index(drop=True).iterrows():
        annotator = row.get("Annotator_ID", "")
        record = row.get("Record_ID", "")
        source = row.get("Sampling_Stratum", "")
        choices.append(f"{index + 1}/{total} | {annotator} | {record} | {source} | {row['Annotation_ID']}")
    return choices


def _review_annotator_choices(frame: pd.DataFrame) -> list[str]:
    if frame.empty or "Annotator_ID" not in frame.columns:
        return []
    annotators = frame["Annotator_ID"].fillna("").astype(str).str.strip()
    return sorted([annotator for annotator in annotators.unique().tolist() if annotator])


def _filter_review_frame(frame: pd.DataFrame, annotator_id: str | None) -> pd.DataFrame:
    if frame.empty or "Annotator_ID" not in frame.columns:
        return frame
    annotator = str(annotator_id or "").strip()
    if not annotator:
        return frame.iloc[0:0].copy()
    return frame[frame["Annotator_ID"].fillna("").astype(str).str.strip().eq(annotator)].copy()


def _annotation_id_from_choice(choice: str | None) -> str:
    if not choice:
        return ""
    return str(choice).split(" | ")[-1].strip()


def _is_filled(value: object) -> bool:
    if pd.isna(value):
        return False
    return str(value).strip() != ""


def _review_progress_html(frame: pd.DataFrame) -> str:
    if frame.empty:
        return '<div class="review-progress">No annotation template loaded yet. Generate Step 2 Template first.</div>'
    required = ["Boundary_Strength_0_3", "Human_C_t", "Human_M_t", "Human_A_t", "Human_R_t"]
    available = [col for col in required if col in frame.columns]
    if not available:
        completed = 0
    else:
        completed = int(frame[available].apply(lambda row: all(_is_filled(value) for value in row), axis=1).sum())
    total = len(frame)
    pct = round(100 * completed / total, 1) if total else 0
    return (
        '<div class="review-progress">'
        f"<strong>Review progress:</strong> {completed}/{total} complete ({pct}%). "
        "A row is complete when boundary strength and four signal labels are filled."
        "</div>"
    )


def _review_completion_table(frame: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Review_Status",
        "Boundary_Strength_0_3",
        "Human_C_t",
        "Human_M_t",
        "Human_A_t",
        "Human_R_t",
        "Human_Primary_Type",
        "Annotator_ID",
        "Record_ID",
        "Sampling_Stratum",
        "Annotation_ID",
    ]
    if frame.empty:
        return pd.DataFrame(columns=columns)

    work = _single_boundary_variable_frame(frame)
    required = ["Boundary_Strength_0_3", "Human_C_t", "Human_M_t", "Human_A_t", "Human_R_t"]
    for col in required:
        if col not in work.columns:
            work[col] = ""
    complete_mask = work[required].apply(lambda row: all(_is_filled(value) for value in row), axis=1)
    work["Review_Status"] = np.where(complete_mask, "complete", "pending")
    for col in columns:
        if col not in work.columns:
            work[col] = ""
    table = work[columns].copy()
    table.insert(0, "Row", range(1, len(table) + 1))
    return table.sort_values(["Review_Status", "Row"], ascending=[False, True]).reset_index(drop=True)


def _value_as_choice(value: object) -> str | None:
    if not _is_filled(value):
        return None
    text = str(value).strip()
    if text in {"0", "0.0", "False", "false"}:
        return "0"
    if text in {"1", "1.0", "True", "true"}:
        return "1"
    return text


def _value_as_bool(value: object) -> bool:
    return _value_as_choice(value) == "1"


def _review_context_html(row: pd.Series | None) -> str:
    if row is None:
        return '<div class="review-context-box">Select an annotation item to begin.</div>'
    context = html.escape(str(row.get("Boundary_Context", "")))
    context = context.replace(
        "[[CANDIDATE_BOUNDARY]]",
        '<span class="review-boundary-marker">CANDIDATE BOUNDARY</span>',
    )
    previous = html.escape(str(row.get("Previous_Segment", "")))
    left = html.escape(str(row.get("Left_Segment", "")))
    right = html.escape(str(row.get("Right_Segment", "")))
    return f"""
<div class="review-context-box">
  <div>{context}</div>
  <div class="review-segment-grid">
    <div class="review-segment-card"><h4>Previous Segment</h4><div>{previous}</div></div>
    <div class="review-segment-card"><h4>Left Segment</h4><div>{left}</div></div>
    <div class="review-segment-card"><h4>Right Segment</h4><div>{right}</div></div>
  </div>
</div>
"""


def _review_metadata_markdown(row: pd.Series | None) -> str:
    if row is None:
        return ""
    fields = [
        ("Annotation_ID", row.get("Annotation_ID", "")),
        ("Annotator_ID", row.get("Annotator_ID", "")),
        ("Record_ID", row.get("Record_ID", "")),
        ("Boundary_Gap", row.get("Boundary_Gap", "")),
        ("Algorithm_Dominant_Signal", row.get("Algorithm_Dominant_Signal", "")),
        ("Algorithm_Selected_Boundary", row.get("Algorithm_Selected_Boundary", "")),
        ("Alg_C_t", row.get("Alg_C_t", "")),
        ("Alg_M_t", row.get("Alg_M_t", "")),
        ("Alg_A_t", row.get("Alg_A_t", "")),
        ("Alg_R_t", row.get("Alg_R_t", "")),
    ]
    lines = ["**Algorithm and item details**"]
    for name, value in fields:
        if _is_filled(value):
            lines.append(f"- `{name}`: {value}")
    return "\n".join(lines)


def _row_for_choice(frame: pd.DataFrame, choice: str | None) -> pd.Series | None:
    annotation_id = _annotation_id_from_choice(choice)
    if frame.empty or not annotation_id or "Annotation_ID" not in frame.columns:
        return None
    matches = frame[frame["Annotation_ID"].astype(str).eq(annotation_id)]
    if matches.empty:
        return None
    return matches.iloc[0]


def _review_item_values(row: pd.Series | None) -> tuple:
    if row is None:
        return (
            _review_context_html(None),
            "",
            None,
            False,
            False,
            False,
            False,
            None,
            "",
            "",
            "",
            "",
            "No annotation item selected.",
        )
    return (
        _review_context_html(row),
        _review_metadata_markdown(row),
        _boundary_strength_from_row(row),
        _value_as_bool(row.get("Human_C_t")),
        _value_as_bool(row.get("Human_M_t")),
        _value_as_bool(row.get("Human_A_t")),
        _value_as_bool(row.get("Human_R_t")),
        None
        if not _is_filled(row.get("Human_Primary_Type", ""))
        else str(row.get("Human_Primary_Type", "")).strip(),
        "" if pd.isna(row.get("Cue_Span", "")) else str(row.get("Cue_Span", "")),
        "" if pd.isna(row.get("Counterexample_or_Exclusion", "")) else str(row.get("Counterexample_or_Exclusion", "")),
        "" if pd.isna(row.get("Codebook_Issue", "")) else str(row.get("Codebook_Issue", "")),
        "" if pd.isna(row.get("Notes", "")) else str(row.get("Notes", "")),
        f"Loaded {row.get('Annotation_ID', '')}.",
    )


def load_review_workspace(annotator_id: str | None = None) -> tuple:
    frame = _review_dataframe()
    annotators = _review_annotator_choices(frame)
    selected_annotator = str(annotator_id or "").strip()
    if selected_annotator not in annotators:
        selected_annotator = annotators[0] if annotators else None
    annotator_frame = _filter_review_frame(frame, selected_annotator)
    choices = _review_choices(annotator_frame)
    selected = choices[0] if choices else None
    row = _row_for_choice(annotator_frame, selected)
    status = (
        f"Loaded review workspace for annotator {selected_annotator}."
        if choices
        else "Generate a Step 2 template first, or choose an annotator with assigned items."
    )
    item_values = _review_item_values(row)
    return (
        gr.update(choices=annotators, value=selected_annotator),
        gr.update(choices=choices, value=selected),
        _review_progress_html(annotator_frame),
        _review_completion_table(annotator_frame),
        *item_values[:-1],
        status,
    )


def select_review_item(choice: str | None) -> tuple:
    frame = _review_dataframe()
    row = _row_for_choice(frame, choice)
    return _review_item_values(row)


def _coerce_binary(value: object) -> str:
    choice = _value_as_choice(value)
    return choice if choice in {"0", "1"} else ""


def _boundary_strength_from_row(row: pd.Series) -> str | None:
    strength = _value_as_choice(row.get("Boundary_Strength_0_3"))
    if strength in {"0", "1", "2", "3"}:
        return strength
    valid = _value_as_choice(row.get("Human_Boundary_Valid"))
    confidence = _value_as_choice(row.get("Confidence_1_3"))
    if valid == "0":
        return "0"
    if valid == "1" and confidence in {"1", "2", "3"}:
        return confidence
    return valid if valid in {"0", "1"} else None


def _save_review_values(
    choice: str | None,
    annotator_id: str | None,
    boundary_strength: str | None,
    human_c: bool,
    human_m: bool,
    human_a: bool,
    human_r: bool,
    primary_type: str | None,
    cue_span: str,
    counterexample: str,
    codebook_issue: str,
    notes: str,
    advance: bool,
) -> tuple:
    frame = _review_dataframe()
    annotator_frame = _filter_review_frame(frame, annotator_id)
    choices = _review_choices(annotator_frame)
    annotation_id = _annotation_id_from_choice(choice)
    if frame.empty or not annotation_id:
        return (
            gr.update(choices=_review_annotator_choices(frame), value=annotator_id),
            gr.update(choices=choices, value=choice),
            _review_progress_html(annotator_frame),
            _review_completion_table(annotator_frame),
            *_review_item_values(None)[:-1],
            "No annotation item selected.",
        )
    mask = frame["Annotation_ID"].astype(str).eq(annotation_id)
    if not mask.any():
        return (
            gr.update(choices=_review_annotator_choices(frame), value=annotator_id),
            gr.update(choices=choices, value=choice),
            _review_progress_html(annotator_frame),
            _review_completion_table(annotator_frame),
            *_review_item_values(None)[:-1],
            f"Annotation_ID not found: {annotation_id}",
        )

    strength_text = str(boundary_strength or "").strip()
    if strength_text not in {"0", "1", "2", "3"}:
        strength_text = ""
    frame.loc[mask, "Boundary_Strength_0_3"] = strength_text
    frame.loc[mask, "Human_C_t"] = "1" if human_c else "0"
    frame.loc[mask, "Human_M_t"] = "1" if human_m else "0"
    frame.loc[mask, "Human_A_t"] = "1" if human_a else "0"
    frame.loc[mask, "Human_R_t"] = "1" if human_r else "0"
    frame.loc[mask, "Human_Primary_Type"] = str(primary_type or "").strip()
    frame.loc[mask, "Cue_Span"] = str(cue_span or "").strip()
    frame.loc[mask, "Counterexample_or_Exclusion"] = str(counterexample or "").strip()
    frame.loc[mask, "Codebook_Issue"] = str(codebook_issue or "").strip()
    frame.loc[mask, "Notes"] = str(notes or "").strip()
    frame = _single_boundary_variable_frame(frame)
    replace_current_table("step2_annotation_items", frame)
    save_current_artifact(
        "step2_in_app_review",
        current_artifact_path("step2_annotation_template") or "",
        {"last_saved_annotation_id": annotation_id},
        f"Saved in-app review labels for {annotation_id}.",
    )

    annotator_frame = _filter_review_frame(frame, annotator_id)
    choices = _review_choices(annotator_frame)
    current_index = choices.index(choice) if choice in choices else 0
    next_index = min(current_index + 1, len(choices) - 1) if advance else current_index
    selected = choices[next_index] if choices else None
    row = _row_for_choice(annotator_frame, selected)
    return (
        gr.update(choices=_review_annotator_choices(frame), value=annotator_id),
        gr.update(choices=choices, value=selected),
        _review_progress_html(annotator_frame),
        _review_completion_table(annotator_frame),
        *_review_item_values(row)[:-1],
        f"Saved {annotation_id} to SQLite.",
    )


def save_review_item(
    choice: str | None,
    annotator_id: str | None,
    boundary_strength: str | None,
    human_c: bool,
    human_m: bool,
    human_a: bool,
    human_r: bool,
    primary_type: str | None,
    cue_span: str,
    counterexample: str,
    codebook_issue: str,
    notes: str,
) -> tuple:
    return _save_review_values(
        choice,
        annotator_id,
        boundary_strength,
        human_c,
        human_m,
        human_a,
        human_r,
        primary_type,
        cue_span,
        counterexample,
        codebook_issue,
        notes,
        advance=False,
    )


def save_review_item_and_next(
    choice: str | None,
    annotator_id: str | None,
    boundary_strength: str | None,
    human_c: bool,
    human_m: bool,
    human_a: bool,
    human_r: bool,
    primary_type: str | None,
    cue_span: str,
    counterexample: str,
    codebook_issue: str,
    notes: str,
) -> tuple:
    return _save_review_values(
        choice,
        annotator_id,
        boundary_strength,
        human_c,
        human_m,
        human_a,
        human_r,
        primary_type,
        cue_span,
        counterexample,
        codebook_issue,
        notes,
        advance=True,
    )


def export_reviewed_template() -> tuple[str | None, str | None, str]:
    frame = _review_dataframe()
    if frame.empty:
        return None, None, "No Step 2 annotation items found. Generate or load a template first."
    frame = _single_boundary_variable_frame(frame)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTPUT_DIR / f"in_app_reviewed_annotation_template_{timestamp}.xlsx"
    instructions_df = pd.DataFrame(ANNOTATION_INSTRUCTIONS)
    codebook = read_current_table("step2_codebook")
    codebook_df = codebook if not codebook.empty else pd.DataFrame(ANNOTATION_CODEBOOK_ROWS)
    short_labels_df = annotation_short_labels_df()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="annotation_items", index=False)
        instructions_df.to_excel(writer, sheet_name="instructions", index=False)
        codebook_df.to_excel(writer, sheet_name="codebook", index=False)
        short_labels_df.to_excel(writer, sheet_name="short_labels", index=False)
        _format_annotation_workbook(writer, frame)
    save_current_artifact("step2_reviewed_annotation_template", out, {"source": "in_app_review"}, "Exported reviewed template from SQLite.")
    log_event("export_reviewed_template", DB_PATH, out, {"source": "step2_annotation_items"}, "success", f"Exported reviewed template: {out}")
    return str(out), str(out), f"Exported reviewed workbook for Step 3: {out}"


def make_annotation_template_and_load_review(
    segmentation_workbook: str | None,
    calibration_round: str,
    sample_size: int,
    annotators: str,
    seed: int,
) -> tuple:
    template_file, template_preview, template_status = make_annotation_template(
        segmentation_workbook,
        calibration_round,
        sample_size,
        annotators,
        seed,
    )
    return (template_file, template_preview, template_status, *load_review_workspace())


def _binary_label(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip().str.lower().map(
        {"1": 1, "true": 1, "yes": 1, "y": 1, "0": 0, "false": 0, "no": 0, "n": 0}
    )


def _boundary_strength_label(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series.fillna("").astype(str).str.strip(), errors="coerce").where(
        lambda values: values.isin([0, 1, 2, 3])
    )


def cohen_kappa(a: Iterable[int], b: Iterable[int]) -> float:
    pairs = [(int(x), int(y)) for x, y in zip(a, b) if pd.notna(x) and pd.notna(y)]
    if not pairs:
        return np.nan
    observed = sum(1 for x, y in pairs if x == y) / len(pairs)
    counts_a = Counter(x for x, _ in pairs)
    counts_b = Counter(y for _, y in pairs)
    labels = sorted(set(counts_a) | set(counts_b))
    expected = sum((counts_a[label] / len(pairs)) * (counts_b[label] / len(pairs)) for label in labels)
    if expected == 1:
        return 1.0 if observed == 1 else np.nan
    return (observed - expected) / (1 - expected)


def evaluate_labels(labeled_workbook: str | None) -> tuple[pd.DataFrame, pd.DataFrame, str | None, str]:
    path = _path_from_upload(labeled_workbook)
    if path is None:
        return pd.DataFrame(), pd.DataFrame(), None, "Upload a completed annotation workbook first."

    try:
        if path.suffix.lower() == ".csv":
            labels = pd.read_csv(path)
        else:
            labels = pd.read_excel(path, sheet_name="annotation_items")
        if "Boundary_Strength_0_3" not in labels.columns and "Human_Boundary_Valid" in labels.columns:
            labels["Boundary_Strength_0_3"] = labels["Human_Boundary_Valid"].fillna("").astype(str).str.strip().map(
                {"0": "0", "1": "2"}
            )
        required = {"Annotation_Item_ID", "Annotator_ID", "Boundary_Strength_0_3", *[f"Human_{signal}" for signal in SIGNALS]}
        missing = required - set(labels.columns)
        if missing:
            raise ValueError(f"labeled file missing columns: {', '.join(sorted(missing))}")

        labels["Boundary_Strength_bin"] = _boundary_strength_label(labels["Boundary_Strength_0_3"])
        for signal in SIGNALS:
            labels[f"{signal}_bin"] = _binary_label(labels[f"Human_{signal}"])

        signal_rows = []
        for signal, col in [("Boundary_Strength_0_3", "Boundary_Strength_bin")] + [
            (signal, f"{signal}_bin") for signal in SIGNALS
        ]:
            complete = labels.dropna(subset=[col])
            item_counts = complete.groupby("Annotation_Item_ID")["Annotator_ID"].nunique()
            multi_items = item_counts[item_counts >= 2].index
            agreement_values = []
            for _item_id, group in complete[complete["Annotation_Item_ID"].isin(multi_items)].groupby("Annotation_Item_ID"):
                values = group.drop_duplicates("Annotator_ID")[col].astype(int).tolist()
                agreement_values.append(1.0 if len(set(values)) == 1 else 0.0)

            kappa = np.nan
            pairable = complete.groupby("Annotation_Item_ID").filter(
                lambda group: group["Annotator_ID"].nunique() == 2
            )
            if not pairable.empty:
                pivot = pairable.pivot_table(index="Annotation_Item_ID", columns="Annotator_ID", values=col, aggfunc="first")
                if pivot.shape[1] == 2:
                    kappa = cohen_kappa(pivot.iloc[:, 0], pivot.iloc[:, 1])

            signal_rows.append(
                {
                    "Signal": signal,
                    "Completed_Labels": int(complete[col].notna().sum()),
                    "Positive_Rate": round(float(complete[col].mean()), 3) if not complete.empty else np.nan,
                    "Multi_Annotated_Items": int(len(multi_items)),
                    "Exact_Agreement": round(float(np.mean(agreement_values)), 3) if agreement_values else np.nan,
                    "Cohen_Kappa_If_2_Annotators": round(float(kappa), 3) if not pd.isna(kappa) else np.nan,
                }
            )

        item_rows = []
        for item_id, group in labels.groupby("Annotation_Item_ID"):
            row = {"Annotation_Item_ID": item_id, "Annotator_Count": group["Annotator_ID"].nunique()}
            boundary_values = group["Boundary_Strength_bin"].dropna().astype(int).tolist()
            row["Boundary_Strength_Values"] = ", ".join(map(str, boundary_values))
            row["Boundary_Strength_Disagreement"] = bool(len(set(boundary_values)) > 1)
            for signal in SIGNALS:
                values = group[f"{signal}_bin"].dropna().astype(int).tolist()
                row[f"{signal}_Positive_Count"] = int(sum(values))
                row[f"{signal}_Disagreement"] = bool(len(set(values)) > 1)
            item_rows.append(row)

        summary = pd.DataFrame(signal_rows)
        item_agreement = pd.DataFrame(item_rows)
        disagreement = item_agreement[
            item_agreement[["Boundary_Strength_Disagreement", *[f"{signal}_Disagreement" for signal in SIGNALS]]].any(axis=1)
        ].copy()
        if not disagreement.empty:
            disagreement = disagreement.merge(
                labels.drop_duplicates("Annotation_Item_ID"),
                on="Annotation_Item_ID",
                how="left",
                suffixes=("", "_example"),
            )

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        out = OUTPUT_DIR / f"{path.stem}_agreement_report.xlsx"
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="signal_agreement", index=False)
            item_agreement.to_excel(writer, sheet_name="item_agreement", index=False)
            disagreement.to_excel(writer, sheet_name="disagreement_items", index=False)

        status = f"Evaluated {len(labels)} label rows across {labels['Annotation_Item_ID'].nunique()} unique items."
        parameters = {"label_rows": len(labels), "unique_items": int(labels["Annotation_Item_ID"].nunique())}
        replace_current_table("step3_completed_labels", labels)
        replace_current_table("step3_signal_agreement", summary)
        replace_current_table("step3_item_agreement", item_agreement)
        replace_current_table("step3_disagreement_items", disagreement)
        save_current_artifact("step3_agreement_report", out, parameters, status)
        log_event(
            "evaluate_labels",
            path,
            out,
            parameters,
            "success",
            status,
        )
        return summary, item_agreement.head(50), str(out), status
    except Exception as exc:
        message = f"Evaluation failed: {exc}"
        log_event("evaluate_labels", path, "", {"workbook": str(path)}, "error", message)
        return pd.DataFrame(), pd.DataFrame(), None, message


def extract_cues(text: object, max_ngram: int = 3) -> set[str]:
    tokens = [token.lower().replace("’", "'") for token in TOKEN_RE.findall(str(text or ""))]
    cues = set()
    for n in range(1, max_ngram + 1):
        for i in range(0, max(0, len(tokens) - n + 1)):
            cues.add(" ".join(tokens[i : i + n]))
    return cues


def analyze_cue_lexicon(
    labeled_workbook: str | None,
    min_support: int,
    min_precision: float,
    min_lift: float,
) -> tuple[str | None, str | None, pd.DataFrame, str]:
    path = _path_from_upload(labeled_workbook)
    if path is None:
        return None, None, pd.DataFrame(), "Upload a completed annotation workbook first."
    try:
        if path.suffix.lower() == ".csv":
            labels = pd.read_csv(path)
        else:
            labels = pd.read_excel(path, sheet_name="annotation_items")
        required = {"Annotation_Item_ID", *[f"Human_{signal}" for signal in SIGNALS]}
        missing = required - set(labels.columns)
        if missing:
            raise ValueError(f"labeled file missing columns: {', '.join(sorted(missing))}")

        context_cols = [col for col in ["Cue_Span", "Boundary_Context", "Left_Segment", "Right_Segment"] if col in labels.columns]
        if not context_cols:
            raise ValueError("labeled file needs at least one text column: Cue_Span, Boundary_Context, Left_Segment, or Right_Segment")

        work = labels.copy()
        work["Cue_Text_For_Analysis"] = work[context_cols].fillna("").astype(str).agg(" ".join, axis=1)
        for signal in SIGNALS:
            work[f"{signal}_bin"] = _binary_label(work[f"Human_{signal}"])

        cue_rows = []
        all_cues_by_row = [extract_cues(text) for text in work["Cue_Text_For_Analysis"]]
        all_cues = sorted(set().union(*all_cues_by_row)) if all_cues_by_row else []
        total_rows = len(work)

        for signal in SIGNALS:
            label_col = f"{signal}_bin"
            valid = work[label_col].notna()
            signal_positive_total = int(work.loc[valid, label_col].sum())
            base_rate = signal_positive_total / int(valid.sum()) if int(valid.sum()) else 0.0
            for cue in all_cues:
                cue_mask = pd.Series([cue in cues for cues in all_cues_by_row], index=work.index) & valid
                support = int(cue_mask.sum())
                if support < int(min_support):
                    continue
                positives = int(work.loc[cue_mask, label_col].sum())
                precision = positives / support if support else 0.0
                recall = positives / signal_positive_total if signal_positive_total else 0.0
                lift = precision / base_rate if base_rate else np.nan
                if precision >= float(min_precision) and (pd.isna(lift) or lift >= float(min_lift)):
                    cue_rows.append(
                        {
                            "Signal": signal,
                            "Cue": cue,
                            "Support": support,
                            "Positive_With_Cue": positives,
                            "Precision": round(precision, 3),
                            "Recall": round(recall, 3),
                            "Lift": round(float(lift), 3) if not pd.isna(lift) else np.nan,
                        }
                    )

        cue_df = pd.DataFrame(cue_rows).sort_values(
            ["Signal", "Precision", "Lift", "Support"],
            ascending=[True, False, False, False],
        ) if cue_rows else pd.DataFrame(columns=["Signal", "Cue", "Support", "Positive_With_Cue", "Precision", "Recall", "Lift"])

        recommended = {}
        for signal in SIGNALS:
            subset = cue_df[cue_df["Signal"].eq(signal)].head(50)
            words = subset[~subset["Cue"].str.contains(" ", regex=False)]["Cue"].tolist()
            phrases = subset[subset["Cue"].str.contains(" ", regex=False)]["Cue"].tolist()
            recommended[signal] = {
                "words": words,
                "phrases": phrases,
                "exclude_patterns": [],
                "criteria": {
                    "min_support": int(min_support),
                    "min_precision": float(min_precision),
                    "min_lift": float(min_lift),
                },
            }

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        report_path = OUTPUT_DIR / f"{path.stem}_cue_lexicon_analysis.xlsx"
        json_path = OUTPUT_DIR / f"{path.stem}_recommended_vocab.json"
        vocab_summary = pd.DataFrame(
            [
                {"Signal": signal, "Recommended_Words": ", ".join(values["words"]), "Recommended_Phrases": ", ".join(values["phrases"])}
                for signal, values in recommended.items()
            ]
        )
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            cue_df.to_excel(writer, sheet_name="recommended_cues", index=False)
            vocab_summary.to_excel(writer, sheet_name="vocab_summary", index=False)
        json_path.write_text(json.dumps(recommended, indent=2, ensure_ascii=False), encoding="utf-8")

        status = f"Analyzed cues from {total_rows} label rows; recommended {len(cue_df)} signal-cue rows."
        parameters = {"min_support": min_support, "min_precision": min_precision, "min_lift": min_lift}
        replace_current_table("step4_labeled_rows_for_cue_analysis", work)
        replace_current_table("step4_recommended_cues", cue_df)
        replace_current_table("step4_vocab_summary", vocab_summary)
        save_current_artifact(
            "step4_cue_lexicon_analysis",
            report_path,
            {**parameters, "recommended_vocab_json": str(json_path)},
            f"{status} JSON: {json_path}",
        )
        log_event(
            "analyze_cue_lexicon",
            path,
            report_path,
            parameters,
            "success",
            f"{status} JSON: {json_path}",
        )
        return str(report_path), str(json_path), cue_df.head(100), status
    except Exception as exc:
        message = f"Cue analysis failed: {exc}"
        log_event(
            "analyze_cue_lexicon",
            path,
            "",
            {"min_support": min_support, "min_precision": min_precision, "min_lift": min_lift},
            "error",
            message,
        )
        return None, None, pd.DataFrame(), message


def build_app() -> gr.Blocks:
    init_db()
    with gr.Blocks(title="DAPS Calibration Lab", css=APP_CSS) as app:
        gr.Markdown("# DAPS Calibration Lab")
        gr.Markdown("Complete workflow: raw transcript segmentation, boundary-centered calibration, cue vocabulary training, and validation logging.")

        with gr.Tab("Step 0 Segment Raw Data"):
            step_header("Step 0. Segment Raw Transcript Data")
            raw_file = gr.File(
                label="Raw transcript CSV/Excel",
                file_types=[".csv", ".xlsx", ".xlsm", ".xls"],
                type="filepath",
            )
            gr.Markdown("For cleaned `data/input.csv`, use defaults: `interview_id` + `text`.")
            with gr.Row():
                raw_sheet_name = gr.Textbox(label="Sheet name", placeholder="Blank = first sheet / CSV ignored")
                raw_id_column = gr.Textbox(label="ID column", value="interview_id")
                raw_item_column = gr.Textbox(label="Item/task column", value="turn_number")
                raw_text_columns = gr.Textbox(label="Transcript column(s)", value="text")
            with gr.Row():
                with gr.Column():
                    help_label("Semantic continuity model")
                    raw_embedding_model = gr.Radio(
                        choices=["lexical", "all-mpnet-base-v2"],
                        value="lexical",
                        show_label=False,
                    )
                with gr.Column():
                    help_label("Clean encoding artifacts / mojibake")
                    raw_clean_text = gr.Checkbox(label="Enable cleanup", value=True)
            with gr.Accordion("Segmentation parameters", open=False):
                with gr.Row(elem_classes=["parameter-row"]):
                    raw_context_width = step0_slider("Context width", "w", 4, 30, 12, 1)
                    raw_local_radius = step0_slider("Local radius", "r", 2, 20, 6, 1)
                    raw_sensitivity = step0_slider("Sensitivity", "tau", 0.05, 1.5, 0.55, 0.05)
                with gr.Row(elem_classes=["parameter-row"]):
                    raw_min_segment_tokens = step0_slider("Minimum segment tokens", "L_min", 3, 40, 12, 1)
                    raw_max_segment_tokens = step0_slider("Maximum segment tokens", "L_max", 30, 120, 60, 5)
                    raw_max_segments_per_record = step0_slider("Maximum segments per record", "K_max", 0, 200, 80, 5)
                    raw_nms_radius = step0_slider("NMS radius", "rho", 1, 15, 6, 1)
            with gr.Accordion("Signal vocabularies", open=False):
                gr.Markdown("Edit words or phrases separated by commas, spaces, or new lines. Step 0 uses these vocabularies immediately when computing DAPS signal scores.")
                help_label("Task-density vocabulary")
                raw_task_vocab = gr.Textbox(
                    value=DEFAULT_VOCABULARY.as_text("task"),
                    lines=3,
                    show_label=False,
                )
                with gr.Row():
                    with gr.Column():
                        help_label("Cognitive transition vocabulary", "V_C")
                        raw_cognitive_vocab = gr.Textbox(
                            value=DEFAULT_VOCABULARY.as_text("cognitive"),
                            lines=5,
                            show_label=False,
                        )
                    with gr.Column():
                        help_label("Metacognitive reset vocabulary", "V_M")
                        raw_metacognitive_vocab = gr.Textbox(
                            value=DEFAULT_VOCABULARY.as_text("metacognitive"),
                            lines=5,
                            show_label=False,
                        )
                with gr.Row():
                    with gr.Column():
                        help_label("Affective friction vocabulary", "V_A")
                        raw_affective_vocab = gr.Textbox(
                            value=DEFAULT_VOCABULARY.as_text("affective"),
                            lines=5,
                            show_label=False,
                        )
                    with gr.Column():
                        help_label("Rhetorical/structural break vocabulary", "V_R")
                        raw_structural_vocab = gr.Textbox(
                            value=DEFAULT_VOCABULARY.as_text("structural"),
                            lines=5,
                            show_label=False,
                        )
            with gr.Row():
                raw_segment_button = gr.Button("Run Initial DAPS Segmentation", variant="primary")
                load_sqlite_button = gr.Button("Load Latest SQLite State", variant="secondary")
            raw_segmented_file = gr.File(label="Download segmented workbook / use this in Step 1 and Step 2")
            raw_segments_preview = gr.Dataframe(label="Segments preview", interactive=False)
            raw_boundaries_preview = gr.Dataframe(label="Boundary evidence preview", interactive=False)
            raw_segmentation_status = gr.Textbox(label="Status", interactive=False)

        with gr.Tab("Step 1 Inspect"):
            step_header("Step 1. Inspect Segmentation Workbook")
            inspect_workbook = gr.File(
                label="DAPS segmentation workbook",
                file_types=[".xlsx"],
                type="filepath",
            )
            inspect_button = gr.Button("Run Pre-Validation Inspection", variant="primary")
            inspection_summary = gr.Dataframe(label="Pre-validation metrics", interactive=False, visible=False)
            inspection_summary_colored = gr.HTML()
            signal_summary = gr.Dataframe(label="Signal saturation summary", interactive=False, visible=False)
            signal_summary_colored = gr.HTML()
            inspection_status = gr.Textbox(label="Status", interactive=False)

        with gr.Tab("Step 2 Template"):
            step_header("Step 2. Generate Multi-Annotator Template")
            gr.Markdown(
                """
**Annotation rule:** use `Boundary_Strength_0_3` for the boundary judgment: `0 = merge`, `1 = weak`, `2 = moderate`, `3 = strong`. Use `0/1` for `Human_C_t`, `Human_M_t`, `Human_A_t`, and `Human_R_t`; the four signal columns are multi-label, so more than one can be `1`. Use `Human_Primary_Type` for the strongest label and `Cue_Span` for the exact evidence phrase. Do not edit ID, context, or algorithm-score columns.
"""
            )
            segmentation_workbook = gr.File(
                label="DAPS segmentation workbook",
                file_types=[".xlsx"],
                type="filepath",
            )
            calibration_round = gr.Radio(
                label="Calibration round",
                choices=["Pilot annotation", "Formal calibration"],
                value="Pilot annotation",
                info="Pilot is usually about 50 items; formal calibration is usually 400-500 items.",
            )
            with gr.Row():
                sample_size = gr.Slider(25, 1000, value=50, step=25, label="Unique boundary items")
                seed = gr.Number(value=20260711, precision=0, label="Random seed")
            annotators = gr.Textbox(
                label="Annotator IDs",
                value="Annotator_A, Annotator_B",
                info="Comma-separated. The same sampled items are duplicated for each annotator.",
            )
            generate_button = gr.Button("Generate Template", variant="primary")
            template_file = gr.File(label="Download annotation template")
            template_preview = gr.Dataframe(label="Template preview", interactive=False)
            template_status = gr.Textbox(label="Status", interactive=False)

        with gr.Tab("Step 2 Review"):
            step_header("Step 2b. In-App Human Review")
            gr.Markdown(
                "Use this guided review screen instead of editing the wide Excel table. Load the latest Step 2 template, review one boundary at a time, save to SQLite, then export the reviewed workbook for Step 3."
            )
            with gr.Row():
                load_review_button = gr.Button("Load Latest Template From SQLite", variant="primary")
                export_review_button = gr.Button("Export Reviewed Workbook", variant="secondary")
            reviewed_template_file = gr.File(label="Download reviewed workbook")
            review_annotator_selector = gr.Dropdown(
                label="Annotator",
                choices=[],
                value=None,
                interactive=True,
                info="Choose the reviewer identity. The item list, progress, and completion table below show only this annotator's rows.",
            )
            review_progress = gr.HTML()
            review_completion = gr.Dataframe(
                label="Review completion table",
                interactive=False,
                wrap=True,
            )
            review_item_selector = gr.Dropdown(
                label="Annotation item",
                choices=[],
                value=None,
                interactive=True,
                info="Select a row to review. The label shows position, annotator, record, stratum, and Annotation_ID.",
            )
            with gr.Row():
                with gr.Column(scale=3):
                    review_context = gr.HTML()
                    review_metadata = gr.Markdown()
                with gr.Column(scale=2):
                    gr.HTML(
                        '<div class="step-title">Boundary Strength'
                        '<span class="help-dot" data-tooltip="Single 0-3 judgment. 0 = merge/not a boundary; 1 = weak or uncertain boundary; 2 = moderate/acceptable boundary; 3 = strong/clear boundary. For analysis, 1-3 are treated as valid boundaries, while 0 is treated as not valid.">?</span></div>'
                    )
                    review_boundary_strength = gr.Radio(
                        label="Boundary",
                        choices=["0", "1", "2", "3"],
                        value=None,
                        interactive=True,
                        info="Required. 0 = merge; 1 = weak; 2 = moderate; 3 = strong.",
                    )
                    gr.HTML(
                        '<div class="step-title">Signal Labels'
                        '<span class="help-dot" data-tooltip="These four labels are independent. More than one can be checked for the same boundary.">?</span></div>'
                    )
                    review_human_c = gr.Checkbox(label="C_t: cognitive")
                    review_human_m = gr.Checkbox(label="M_t: metacognitive")
                    review_human_a = gr.Checkbox(label="A_t: affective")
                    review_human_r = gr.Checkbox(label="R_t: structural")
                    gr.HTML(
                        '<div class="step-title">Primary Type'
                        '<span class="help-dot" data-tooltip="Choose the strongest single explanation. Use mixed when multiple signals are equally central; unclear when you cannot decide.">?</span></div>'
                    )
                    review_primary = gr.Radio(
                        label="Primary",
                        choices=REVIEW_PRIMARY_TYPES,
                        value=None,
                        interactive=True,
                        info="Recommended. Choose the strongest signal type.",
                    )
            with gr.Row():
                with gr.Column():
                    gr.HTML(
                        '<div class="step-title">Evidence Log: 1. Cue'
                        '<span class="help-dot" data-tooltip="This is Cue_Span in the exported data. Copy the exact transcript words that support a signal label you marked as 1. Most valid boundaries should have a short supporting cue.">?</span></div>'
                    )
                    review_cue_span = gr.Textbox(
                        label="Cue (recommended if signal = 1)",
                        lines=3,
                        placeholder="What made you mark a signal as 1? Examples: wait, no; actually; flip it over; then; [Pause]",
                    )
                with gr.Column():
                    gr.HTML(
                        '<div class="step-title">Evidence Log: 2. Reject (optional)'
                        '<span class="help-dot" data-tooltip="This is Counterexample_or_Exclusion in the exported data. Use only when a word looks like a cue but you decide it should not count in this context.">?</span></div>'
                    )
                    review_counterexample = gr.Textbox(
                        label="Reject (optional)",
                        lines=3,
                        placeholder="What did you deliberately NOT count? Examples: looks like = visual comparison, not A_t; no = answer choice label, not M_t",
                    )
            gr.HTML(
                '<div class="step-title">Evidence Log: 3. Issue (optional)'
                '<span class="help-dot" data-tooltip="This is Codebook_Issue in the exported data. Use only when this case reveals that the annotation rule or vocabulary needs clarification.">?</span></div>'
            )
            review_codebook_issue = gr.Textbox(
                label="Issue (optional)",
                lines=3,
                placeholder="What should the team discuss later? Examples: wait is M_t or R_t here? Need rule for interviewer prompts. Boundary valid but no signal definition fits.",
            )
            review_notes = gr.Textbox(label="Notes (optional)", lines=2)
            with gr.Row():
                save_review_button = gr.Button("Save Current", variant="secondary")
                save_next_review_button = gr.Button("Save and Next", variant="primary")
            review_status = gr.Textbox(label="Review status", interactive=False)

        with gr.Tab("Step 3 Agreement"):
            step_header("Step 3. Evaluate Multi-Annotator Labels")
            labeled_workbook = gr.File(
                label="Completed annotation workbook or CSV",
                file_types=[".xlsx", ".csv"],
                type="filepath",
            )
            evaluate_button = gr.Button("Evaluate Multi-Annotator Labels", variant="primary")
            agreement_summary = gr.Dataframe(label="Signal agreement summary", interactive=False)
            item_summary = gr.Dataframe(label="Item disagreement preview", interactive=False)
            agreement_report_file = gr.File(label="Download agreement report")
            evaluation_status = gr.Textbox(label="Status", interactive=False)

        with gr.Tab("Step 4 Cue Lexicon"):
            step_header("Step 4. Analyze Cue Lexicon")
            cue_labeled_workbook = gr.File(
                label="Completed annotation workbook or CSV",
                file_types=[".xlsx", ".csv"],
                type="filepath",
            )
            with gr.Row():
                min_support = gr.Slider(1, 30, value=5, step=1, label="Minimum cue support")
                min_precision = gr.Slider(0.30, 1.00, value=0.65, step=0.05, label="Minimum precision")
                min_lift = gr.Slider(1.00, 5.00, value=1.50, step=0.10, label="Minimum lift")
            cue_button = gr.Button("Analyze Cues and Export Vocabulary", variant="primary")
            cue_report_file = gr.File(label="Download cue analysis workbook")
            vocab_json_file = gr.File(label="Download recommended vocabulary JSON")
            cue_preview = gr.Dataframe(label="Recommended cue preview", interactive=False)
            cue_status = gr.Textbox(label="Status", interactive=False)

        with gr.Tab("Step 5 History"):
            step_header("Step 5. SQLite History")
            history_limit = gr.Slider(10, 500, value=100, step=10, label="Rows to show")
            refresh_history = gr.Button("Refresh SQLite History", variant="primary")
            current_artifacts_table = gr.Dataframe(label="Current latest artifacts and parameters", interactive=False)
            history_table = gr.Dataframe(label=f"Append-only event history: {DB_PATH}", interactive=False)

        raw_segment_button.click(
            run_initial_segmentation_and_inspect,
            inputs=[
                raw_file,
                raw_sheet_name,
                raw_id_column,
                raw_item_column,
                raw_text_columns,
                raw_embedding_model,
                raw_clean_text,
                raw_context_width,
                raw_local_radius,
                raw_sensitivity,
                raw_min_segment_tokens,
                raw_max_segment_tokens,
                raw_max_segments_per_record,
                raw_nms_radius,
                raw_task_vocab,
                raw_cognitive_vocab,
                raw_metacognitive_vocab,
                raw_affective_vocab,
                raw_structural_vocab,
            ],
            outputs=[
                raw_segmented_file,
                raw_segments_preview,
                raw_boundaries_preview,
                raw_segmentation_status,
                inspect_workbook,
                segmentation_workbook,
                inspection_summary,
                signal_summary,
                inspection_summary_colored,
                signal_summary_colored,
                inspection_status,
            ],
        )
        inspect_button.click(
            inspect_segmentation_workbook_for_ui,
            inputs=[inspect_workbook],
            outputs=[
                inspection_summary,
                signal_summary,
                inspection_summary_colored,
                signal_summary_colored,
                inspection_status,
            ],
        )
        review_workspace_outputs = [
            review_annotator_selector,
            review_item_selector,
            review_progress,
            review_completion,
            review_context,
            review_metadata,
            review_boundary_strength,
            review_human_c,
            review_human_m,
            review_human_a,
            review_human_r,
            review_primary,
            review_cue_span,
            review_counterexample,
            review_codebook_issue,
            review_notes,
            review_status,
        ]
        review_item_outputs = [
            review_context,
            review_metadata,
            review_boundary_strength,
            review_human_c,
            review_human_m,
            review_human_a,
            review_human_r,
            review_primary,
            review_cue_span,
            review_counterexample,
            review_codebook_issue,
            review_notes,
            review_status,
        ]
        review_save_inputs = [
            review_item_selector,
            review_annotator_selector,
            review_boundary_strength,
            review_human_c,
            review_human_m,
            review_human_a,
            review_human_r,
            review_primary,
            review_cue_span,
            review_counterexample,
            review_codebook_issue,
            review_notes,
        ]
        generate_button.click(
            make_annotation_template_and_load_review,
            inputs=[segmentation_workbook, calibration_round, sample_size, annotators, seed],
            outputs=[template_file, template_preview, template_status, *review_workspace_outputs],
        )
        load_review_button.click(
            load_review_workspace,
            inputs=[review_annotator_selector],
            outputs=review_workspace_outputs,
        )
        review_annotator_selector.change(
            load_review_workspace,
            inputs=[review_annotator_selector],
            outputs=review_workspace_outputs,
        )
        review_item_selector.change(
            select_review_item,
            inputs=[review_item_selector],
            outputs=review_item_outputs,
        )
        save_review_button.click(
            save_review_item,
            inputs=review_save_inputs,
            outputs=review_workspace_outputs,
        )
        save_next_review_button.click(
            save_review_item_and_next,
            inputs=review_save_inputs,
            outputs=review_workspace_outputs,
        )
        export_review_button.click(
            export_reviewed_template,
            inputs=[],
            outputs=[reviewed_template_file, labeled_workbook, review_status],
        )
        evaluate_button.click(
            evaluate_labels,
            inputs=[labeled_workbook],
            outputs=[agreement_summary, item_summary, agreement_report_file, evaluation_status],
        )
        cue_button.click(
            analyze_cue_lexicon,
            inputs=[cue_labeled_workbook, min_support, min_precision, min_lift],
            outputs=[cue_report_file, vocab_json_file, cue_preview, cue_status],
        )
        refresh_history.click(
            refresh_sqlite_state,
            inputs=[history_limit],
            outputs=[history_table, current_artifacts_table],
        )
        sqlite_restore_outputs = [
            raw_segmented_file,
            raw_segments_preview,
            raw_boundaries_preview,
            raw_segmentation_status,
            inspect_workbook,
            segmentation_workbook,
            inspection_summary,
            signal_summary,
            inspection_summary_colored,
            signal_summary_colored,
            inspection_status,
            history_table,
            current_artifacts_table,
        ]
        load_sqlite_button.click(
            load_latest_sqlite_state,
            inputs=[],
            outputs=sqlite_restore_outputs,
        )
        app.load(
            load_latest_sqlite_state,
            inputs=[],
            outputs=sqlite_restore_outputs,
        )

    return app


if __name__ == "__main__":
    demo = build_app()
    port = int(os.environ.get("PORT", os.environ.get("GRADIO_SERVER_PORT", "8000")))
    server_name = os.environ.get("GRADIO_SERVER_NAME", "0.0.0.0")
    demo.launch(server_name=server_name, server_port=port)
